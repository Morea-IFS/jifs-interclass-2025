from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse, QueryDict
from .models import Sexo_types, Settings_access, UserSession, Event_unit, Event_badge, Detailed, Status, Authenticity, Match_referee, Type_referee, Replacement, Group_phase, Phase, Phase_types, Campus_types, Help, Type_penalties, Detailed, Activity, Statement, Point_types, Event, Event_sport, Statement_user, Users_types, Type_service, Certificate, Attachments, Volley_match, Player, Sport_types, Voluntary, Penalties, Occurrence, Time_pause, Team, Point, Team_sport, Player_team_sport, Match, Team_match, Player_match, Assistance,  Banner, Terms_Use
from django.db.models import Count, Q, Prefetch
from .decorators import time_restriction
from django.contrib import messages
from django.db import IntegrityError
from django.templatetags.static import static
from django.contrib.auth.models import User, Group, Permission
from django.contrib.sessions.models import Session
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth import login as auth_login, authenticate, logout, get_user_model
from django.template.loader import render_to_string
from .forms import Terms_UseForm
from datetime import date, datetime
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from .generators import generate_certificates, generate_badges, generate_events, generate_timer
import time, pytz, os, random
from django.core.files.base import ContentFile
from weasyprint import HTML
from django.utils import timezone
from .decorators import terms_accept_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.urls import reverse
from django.contrib.auth import update_session_auth_hash
from django.core.exceptions import PermissionDenied
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import pytz
from datetime import datetime as dt
from collections import Counter
from django.db.models.functions import TruncDate
from django.db.models import Count
from datetime import date, timedelta
from datetime import date, timedelta
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, permission_required
from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from .models import AccessLog, UserSession
from .decorators import terms_accept_required
from datetime import date, timedelta
from .models import AccessLog, Event

User = get_user_model()

# ─── helpers de gênero ──────────────────────────────────────────────────────

SEXO_NAMES = {0: "masculino", 1: "feminino", 2: "misto"}

def _check_gender_compatibility(player, team_sport, user):
    """
    Verifica se o jogador pode entrar no team_sport.
    Retorna (ok: bool, erro: str | None).
    Superusuários passam direto.
    """
    if user.is_superuser:
        return True, None

    team_sexo = team_sport.sexo  # 0 masc | 1 fem | 2 misto

    # time misto aceita qualquer sexo
    if team_sexo == 2:
        return True, None

    # se o jogador já tem sexo definido, ele deve bater com o time
    if player.sexo is not None and player.sexo != team_sexo:
        return False, (
            f"O atleta '{player.name}' é do sexo "
            f"{SEXO_NAMES.get(player.sexo, '?')} e não pode ser adicionado "
            f"a um time {SEXO_NAMES.get(team_sexo, '?')}."
        )

    # verifica vínculos existentes: se o jogador já está num time
    # de sexo diferente (e não misto), bloqueia
    conflito = (
        Player_team_sport.objects
        .filter(player=player)
        .exclude(team_sport=team_sport)
        .exclude(team_sport__sexo=2)          # ignora times mistos
        .exclude(team_sport__sexo=team_sexo)  # ignora times do mesmo sexo
        .select_related("team_sport__team")
        .first()
    )
    if conflito:
        return False, (
            f"O atleta '{player.name}' já participa do time "
            f"'{conflito.team_sport.team.name}' "
            f"({SEXO_NAMES.get(conflito.team_sport.sexo, '?')}) "
            f"e não pode entrar num time {SEXO_NAMES.get(team_sexo, '?')}."
        )

    return True, None


def _player_queryset_for_team(team_sport, user, event, admin_user=None):
    """
    Retorna queryset de Player compatível com o sexo do team_sport.
    admin_user restringe por quem cadastrou (para tipo 2).
    """
    qs = Player.objects.filter(event=event)

    if admin_user:
        qs = qs.filter(admin=admin_user)

    sexo = team_sport.sexo
    if sexo in (0, 1) and not user.is_superuser:
        # exclui jogadores com sexo INCOMPATÍVEL (nulo é permitido ainda)
        qs = qs.exclude(sexo__in=[s for s in (0, 1) if s != sexo])
        # exclui jogadores já vinculados a times de sexo oposto
        qs = qs.exclude(
            player_team_sport__team_sport__sexo__in=[s for s in (0, 1) if s != sexo]
        )

    return qs.distinct()

def events_list(request):
    if request.method == 'GET':
        events = Event.objects.all().order_by('-id')
        if len(events) == 1:
            return redirect('home_public', events[0].id)
        return render(request, 'public/events_list.html', {'events': events})
    else:
        return redirect('events')

def has_accepted_terms(user, request):
    try:
        user_accepted = User.objects.get(id=request.user.id)
        return bool(user_accepted.document and user_accepted.photo and user_accepted.accepted)
    except Terms_Use.DoesNotExist:
        return False
        
def page_in_erro404(request):
    return render(request, 'error_404.html', status=404)

def about_us(request, event_id):
    event = Event.objects.get(id=event_id)
    return render(request, 'public/about_us.html',{'event':event})

@login_required(login_url="login")
@permission_required('app.view_event', raise_exception=True)
def event_manage(request):
    if request.method == "GET":
        events = Event.objects.all()
        events_unit = Event.objects.filter(general_need_unit=True)
        return render(request, 'events/event_manage.html', {
            'events': events,
            'events_unit': events_unit,
            'sports': Sport_types.choices
        })

    else:
        print(request.POST, request.FILES)
        if 'event' in request.POST:
            event_id = request.POST.get('event')
            sport = request.POST.get('sport')
            min_sport = request.POST.get('min_sport')
            max_sport = request.POST.get('max_sport')
            fem = 'fem' in request.POST
            masc = 'masc' in request.POST
            mist = 'mist' in request.POST

            Event_sport.objects.create(
                event=Event.objects.get(id=event_id),
                sport=sport,
                min_sport=min_sport,
                max_sport=max_sport,
                fem=fem,
                masc=masc,
                mist=mist,
            )
        
        elif 'event_badge' in request.POST:
            event_badge = request.POST.get("event_badge")
            type_badge = request.POST.get("type_badge")
            image_badge = request.FILES.get("image_badge")
            if image_badge and type_badge and image_badge:
                if Event_badge.objects.filter(event=Event.objects.get(id=event_badge), number=type_badge):
                    event_badge = Event_badge.objects.filter(event=Event.objects.get(id=event_badge), number=type_badge)[0]
                    event_badge.file = image_badge
                    event_badge.save()
                    messages.success(request, f"Sucesso! o crachá já existe. sendo assim, apenas a imagem foi trocada.")
                else:
                    Event_badge.objects.create(event=Event.objects.get(id=event_badge), number=type_badge, file=image_badge)
                    messages.success(request, f"Sucesso! o crachá foi cadastrado!.")
            else:
                messages.error(request, f"Erro! algumas informações não foram enviadas. :(")
                

        elif 'event_unit' in request.POST:
            print(request.POST)
            event_unit = int(request.POST.get('event_unit'))
            event=Event.objects.get(id=event_unit)
            name_unit = request.POST.get('name_unit').replace(" ", "").replace(".", "").replace("-", "")
            create_team_user = 'create-team-user' in request.POST
            if Event_unit.objects.filter(event=event, name=name_unit):
                event_unit = Event_unit.objects.get(event=event, name=name_unit)
            else:
                event_unit = Event_unit.objects.create(
                    event=event,
                    name=name_unit,
                )
            event_unit.save()
            if create_team_user:
                if Team.objects.filter(name=name_unit, unit=event_unit, event=event):
                    if User.objects.filter(username=f"admin.{name_unit.lower()}{event.id}", type=2, team=Team.objects.get(name=name_unit, unit=event_unit, event=event), event_user=event, unit=event_unit):
                        messages.info(request, f"Erro! Já existe um time e um usuário que está associada a unidade.")
                        return redirect('event_manage')
                team = Team.objects.create(name=name_unit, unit=event_unit, event=event, status=False)
                team.save()
                list = [] 
                possibilidade = ["0","1","2","3","4","5","6","7","8","9","A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
                list.append(''.join(random.sample(possibilidade, 8)))
                result = str(''.join(list))
                user = User.objects.create_user(username=f"admin.{name_unit.lower()}{event.id}", password=result, type=2, team=team, event_user=event, unit=event_unit)
                messages.success(request, f"Sucesso! Unidade cadastrada com exito.")
                
                cont = {
                    'now': timezone.now(),
                    'user': request.user,
                    'logo_atum': request.build_absolute_uri('/static/images/logo_atum.png')
                }
                
                if event:
                    cont['logo_ifs'] = request.build_absolute_uri(event.logo.url)
                    cont['event'] = event
                else:
                    cont['logo_ifs'] = request.build_absolute_uri('/static/images/logo-jiifs-2025.jpg')
                logo_ifs = cont['logo_ifs']
                cont['logo_ifs_ofc'] = request.build_absolute_uri('/static/images/logo-ifs.svg')
                cont['logo_morea'] = request.build_absolute_uri('/static/images/logo-morea-sports.svg')
                link = f"https://{request.get_host()}"
                qr = qrcode.make(link)
                print(link)
                buffer = BytesIO()
                qr.save(buffer, format='PNG')
                buffer.seek(0)
                cont['site'] = link

                img_base64 = base64.b64encode(buffer.getvalue()).decode()

                img = ImageReader(logo_ifs)
                largura, altura = img.getSize()

                if largura == altura:
                    print("Quadrada")
                    cont['logo_event_type'] = 0
                elif largura > altura:
                    print("Retangular horizontal")
                    cont['logo_event_type'] = 1
                else:
                    print("Retangular vertical")
                    cont['logo_event_type'] = 2

                name_html = 'data-welcome'
                name_pdf = f'boas vindas, {name_unit.lower()}'

                cont['name'] = user.username
                cont['qrcode'] = img_base64
                cont['password'] = result
            
                html_string = render_to_string(f'generator/{name_html}.html', cont)

                response = HttpResponse(content_type='application/pdf')
                #response['Content-Disposition'] = f'inline; filename="{name_pdf}.pdf"'  #gerar e deixar no navegador
                response['Content-Disposition'] = f'attachment; filename="{name_pdf}.pdf"' #gerar e baixar no navegador
                HTML(string=html_string).write_pdf(response)

                return response

        elif 'name' in request.POST:
            name = request.POST.get('name')
            logo = request.FILES.get('logo')
            logo_badge = request.FILES.get('logo_badge')
            description = request.POST.get('description')
            date_init = request.POST.get('date_init')
            date_end = request.POST.get('date_end')
            enrollment_init = request.POST.get('enrollment_init')
            enrollment_end = request.POST.get('enrollment_end')
            local = request.POST.get('local')
            age = request.POST.get('age')
            age_max = request.POST.get('age_max')
            regulation = request.FILES.get('regulation')

            player_need_instagram = 'player_need_instagram' in request.POST
            player_need_photo = 'player_need_photo' in request.POST
            player_need_bulletin = 'player_need_bulletin' in request.POST
            player_need_rg = 'player_need_rg' in request.POST
            player_need_sexo = 'player_need_sexo' in request.POST
            player_need_registration = 'player_need_registration' in request.POST
            player_need_cpf = 'player_need_cpf' in request.POST
            player_need_date_nasc = 'player_need_date_nasc' in request.POST
            player_need_address = 'player_need_address' in request.POST
            player_need_photo_goal = 'player_need_photo_goal' in request.POST
            player_need_course = 'player_need_course' in request.POST
            player_need_cep = 'player_need_cep' in request.POST
            player_need_municipality = 'player_need_municipality' in request.POST

            general_need_authorization = 'general_need_authorization' in request.POST
            general_need_terms = 'general_need_terms' in request.POST
            general_need_unit = 'general_need_unit' in request.POST

            team_need_description = 'team_need_description' in request.POST
            team_need_color = 'team_need_color' in request.POST
            team_need_technician = 'team_need_technician' in request.POST

            terms_intro_text = request.POST.get('terms_intro_text', '').strip() or None
            terms_declaration_text = request.POST.get('terms_declaration_text', '').strip() or None
            upload_intro_text = request.POST.get('upload_intro_text', '').strip() or None

            Event.objects.create(
                name=name,
                logo=logo,
                logo_badge=logo_badge,
                description=description,
                date_init=date_init,
                date_end=date_end,
                enrollment_init=enrollment_init,
                enrollment_end=enrollment_end,
                local=local,
                age=age,
                age_max=age_max,
                regulation=regulation,
                user=request.user,

                player_need_instagram=player_need_instagram,
                player_need_photo=player_need_photo,
                player_need_bulletin=player_need_bulletin,
                player_need_rg=player_need_rg,
                player_need_sexo=player_need_sexo,
                player_need_registration=player_need_registration,
                player_need_cpf=player_need_cpf,
                player_need_date_nasc=player_need_date_nasc,
                player_need_address=player_need_address,
                player_need_photo_goal=player_need_photo_goal,
                player_need_course=player_need_course,
                player_need_cep=player_need_cep,
                player_need_municipality=player_need_municipality,

                general_need_authorization=general_need_authorization,
                general_need_terms=general_need_terms,
                general_need_unit=general_need_unit,

                team_need_description=team_need_description,
                team_need_color=team_need_color,
                team_need_technician=team_need_technician,
                
                terms_intro_text=terms_intro_text,
                terms_declaration_text=terms_declaration_text,
                upload_intro_text=upload_intro_text,
            )

        return redirect('event_manage')

@login_required(login_url="login")
def event_sport_manage(request):
    return render(request, 'events/event_manage.html')

@login_required(login_url="login")
def event_sport_edit(request):
    return render(request, 'events/event_manage.html')

def home_public(request, event_id):
        event = Event.objects.get(id=event_id)
        hoje = timezone.localdate()
        games_day = Match.objects.filter(time_match__date=hoje, event=event).prefetch_related('teams__team').order_by('time_match')
        context_games_day = [
            {
                'match': match,
                'times': list(match.teams.all()),
            }
            for match in games_day
        ]


        volei_masc = Volley_match.objects.filter(matches__sexo=0, event=event).prefetch_related('matches__teams__team').distinct()
        context_volei_masc = [
            {
                'volley_match': volley_match,
                'sets_team_a': volley_match.sets_team_a,
                'sets_team_b': volley_match.sets_team_b,
                'matches': [
                    {
                        'match': match,
                        'times': [
                            {
                                'team': team_match.team,
                                'name': team_match.team.name,
                                'photo_url': team_match.team.photo.url,
                                'points': Point.objects.filter(team_match=team_match).count()
                            }
                            for team_match in match.teams.all()
                        ]
                    }
                    for match in volley_match.matches.all().order_by('time_match')
                ]
            }
            for volley_match in volei_masc
        ]

        volei_fem = Volley_match.objects.filter(matches__sexo=1, event=event).prefetch_related('matches__teams__team').distinct()
        context_volei_fem = [
            {
                'volley_match': volley_match,
                'sets_team_a': volley_match.sets_team_a,
                'sets_team_b': volley_match.sets_team_b,
                'matches': [
                    {
                        'match': match,
                        'times': [
                            {
                                'team': team_match.team,
                                'name': team_match.team.name,
                                'photo_url': team_match.team.photo.url,
                                'points': Point.objects.filter(team_match=team_match).count()
                            }
                            for team_match in match.teams.all()
                        ]
                    }
                    for match in volley_match.matches.all().order_by('time_match')
                ]
            }
            for volley_match in volei_fem
        ]
        
        matchs_futsal_masc = Match.objects.filter(sport=0, sexo=0, event=event).prefetch_related('teams__team').order_by('time_match')
        context_futsal_masc = [
            {
                'match': match,
                'times': list(match.teams.all()),
                'points_a': Point.objects.filter(team_match=match.teams.first()).count(),
                'points_b': Point.objects.filter(team_match=match.teams.last()).count(),
            }
            for match in matchs_futsal_masc
        ]

        matchs_futsal_fem = Match.objects.filter(sport=0, sexo=1, event=event).prefetch_related('teams__team').order_by('time_match')
        context_futsal_fem = [
            {
                'match': match,
                'times': list(match.teams.all()),
                'points_a': Point.objects.filter(team_match=match.teams.first()).count(),
                'points_b': Point.objects.filter(team_match=match.teams.last()).count(),
            }
            for match in matchs_futsal_fem

        ]

        matchs_handebol_masc = Match.objects.filter(sport=3, sexo=0, event=event).prefetch_related('teams__team').order_by('time_match')
        context_handebol_masc = [
            {
                'match': match,
                'times': list(match.teams.all()),
                'points_a': Point.objects.filter(team_match=match.teams.first()).count(),
                'points_b': Point.objects.filter(team_match=match.teams.last()).count(),
            }
            for match in matchs_handebol_masc
        ]

        matchs_handebol_fem = Match.objects.filter(sport=3, sexo=1, event=event).prefetch_related('teams__team').order_by('time_match')
        context_handebol_fem = [
            {
                'match': match,
                'times': list(match.teams.all()),
                'points_a': Point.objects.filter(team_match=match.teams.first()).count(),
                'points_b': Point.objects.filter(team_match=match.teams.last()).count(),
            }
            for match in matchs_handebol_fem

        ]

        matchs_queimado_fem = Match.objects.filter(sport=8, sexo=0, event=event).prefetch_related('teams__team').order_by('time_match')
        context_queimado_fem = [
            {
                'match': match,
                'times': list(match.teams.all()),
                'points_a': Point.objects.filter(team_match=match.teams.first()).count(),
                'points_b': Point.objects.filter(team_match=match.teams.last()).count(),
            }
            for match in matchs_queimado_fem
        ]

        matchs_queimado_masc = Match.objects.filter(sport=8, sexo=1, event=event).prefetch_related('teams__team').order_by('time_match')
        context_queimado_masc = [
            {
                'match': match,
                'times': list(match.teams.all()),
                'points_a': Point.objects.filter(team_match=match.teams.first()).count(),
                'points_b': Point.objects.filter(team_match=match.teams.last()).count(),
            }
            for match in matchs_queimado_masc
        ]
        event_sports = Event_sport.objects.filter(event=event)
        attachments = Attachments.objects.filter(public=True, event=event)

        if request.method == "GET":
            context = {
                'context_queimado_masc':context_queimado_masc,
                'context_queimado_fem':context_queimado_fem,
                'context_volei_masc':context_volei_masc,
                'context_volei_fem':context_volei_fem,
                'context_futsal_masc':context_futsal_masc,
                'context_futsal_fem':context_futsal_fem,
                'context_handebol_masc':context_handebol_masc,
                'context_handebol_fem':context_handebol_fem,
                'context_games_day':context_games_day,
                'event':event,
                'event_sports':event_sports,
                'attachments':attachments,
                'Phase_types': Phase_types,
            }
            
            print(context)
            return render(request, 'public/home_public.html', context)
        
def switching_public(request, event_id):
        print(request.GET)
        event = Event.objects.get(id=event_id)
        event_sports = Event_sport.objects.filter(event=event)

        if request.method == "GET":
            context = {
                'event':event,
                'event_sports':event_sports,
                'phase_types': Phase_types.choices,
                'sexo_types': Sexo_types.choices,
                'phases_all': Phase.objects.filter(event__event=event).order_by('event'),
            }
            phases = Phase.objects.filter(event__event=event)\
                .prefetch_related(
                    'groups__group_matches__teams__team', 
                )
            if request.GET.get('sport') and request.GET.get('sport') != '':
                event_sport = get_object_or_404(Event_sport, id=request.GET.get('sport'))
                phases = phases.filter(event=event_sport)
                context['event_sport'] = event_sport

            if request.GET.get('genre') and request.GET.get('genre') != '':
                phases = phases.filter(sexo=int(request.GET.get('genre')))
                context['genre'] = int(request.GET.get('genre'))

            if request.GET.get('phase') and request.GET.get('phase') != '':
                phases = phases.filter(name=int(request.GET.get('phase')))
                context['phase'] = int(request.GET.get('phase'))

            matches_by_phase = {}

            for phase in phases:
                matches_no_group = Match.objects.filter(
                    event=event,
                    group_phase__isnull=True,
                )

                matches_in_groups = Match.objects.filter(
                    event=event,
                    group_phase__phase=phase,
                )

                matches_by_phase[phase.id] = matches_no_group | matches_in_groups
                matches_by_phase[phase.id] = matches_by_phase[phase.id].prefetch_related('teams__team')

            
            context['phases'] = phases
            context['matches_by_phase'] = matches_by_phase
            
            print(context)
            return render(request, 'public/switching.html', context)
        else:
            return redirect('switching')

@login_required(login_url="login")
@terms_accept_required
def home_admin(request):
    user = request.user
    if user.type != 0:
        event = Event.objects.get(id=user.event_user.id)
        help = Help.objects.all()
        ins = Settings_access.objects.all().last()
        vistos = Statement_user.objects.filter(user=user).values_list('statement_id', flat=True)

        statements_faltando = Statement.objects.exclude(id__in=vistos)

        if statements_faltando.exists():
            imagem_filter = statements_faltando.first()
            imagem = Statement.objects.get(id=imagem_filter.id)
            Statement_user.objects.create(user=user, statement=imagem)
        else:
            imagem = None
    
        return render(request, 'home_admin.html',{'help':help,'ins':ins,'mensagem':'mensagem','imagem':imagem,'event':event})
    else:
        event = Event.objects.all().order_by('active')
        context = {
            'event':event,
            'total_events': Event.objects.all().count(),
            'total_teams': Team_sport.objects.all().count(),
            'total_players': Player.objects.all().count(),
            'total_users': User.objects.all().count(),
        }
        return render(request, 'home_admin_adm.html', context)

def login(request):
        if request.user.is_authenticated == False:
            if request.method == "GET":
                return render(request, 'login.html')
            else:
                username = request.POST.get('username')
                password = request.POST.get('password')
                user = authenticate(username=username, password=password)
                if user:
                    auth_login(request, user)
                    if user.team:
                        messages.success(request, f"Seja bem-vindo time {user.team.name}! para navegar, acesse o menu.")
                    else:
                        messages.success(request, f"Seja bem-vindo ao sistema {user.username}!")
                    next_url = request.GET.get('next') or '/morea-admin'
                    print(next_url)
                    return redirect(next_url)
                else:
                    messages.error(request,"Poxa! algo está errado, pode ser o usuário ou a senha.")
                    return redirect('login')
        else:
            next_url = request.GET.get('next') or '/morea-admin'
            print(next_url)
            return redirect(next_url)

@login_required(login_url="login")
@terms_accept_required
def sair(request):
    logout(request)
    return redirect('events')

@login_required(login_url="login")
@terms_accept_required
def attachments(request):
    if request.method == "GET":
        context = {}
        if request.user.type == 0 or request.user.is_staff:  
            context['events'] = Event.objects.all()
        if request.user.type != 0:
            context['attachments'] = Attachments.objects.filter(event=request.user.event_user)
        elif 'e' in request.GET and request.GET.get('e') != '':
            context['attachments'] = Attachments.objects.filter(event__id=request.GET.get('e'))
            context['select_event'] = request.GET.get('e')

        return render(request, 'attachments.html', context)
    else:
        name = request.POST.get('name')
        public = 'public' in request.POST
        file = request.FILES.get('file')
        if request.user.type == 0:
            Attachments.objects.create(name=name, file=file, event=Event.objects.get(id=request.POST.get('event')), public=public)
        else:
            Attachments.objects.create(name=name, file=file, event=request.user.event_user, public=public)
        return redirect("attachments")

def erro_403_customizado(request, exception=None):
    messages.info(request, "Você não tem permissão para acessar essa página. Contate o administrador.")
    return redirect('Home')

def erro_404_customizado(request, exception):
    return render(request, 'public/error_404.html', status=404)

@login_required(login_url="login")
@permission_required('app.view_player', raise_exception=True)
@terms_accept_required
def player_manage(request):
    if request.method == "GET":
        context = {
            'events': Event.objects.all()
        }
        if request.user.type == 1:
            player = Player.objects.filter(event=request.user.event_user).order_by('id')
        elif request.user.type == 2:
            player = Player.objects.filter(event=request.user.event_user, admin=request.user).order_by('id')
        elif 'e' in request.GET and request.GET['e'] != '' and request.GET['e'] != '0':
            event = Event.objects.get(id=request.GET['e'])
            player = Player.objects.filter(event=event).order_by('id')
            context['select_event'] = request.GET['e']
        else:
            if request.user.is_staff:  
                player = Player.objects.all().order_by('-id')
            else:
                player = Player.objects.filter(admin=request.user).order_by('-id')
        page = request.GET.get('page', 1) 
        paginator = Paginator(player, 20) 
        print(player)
        try:
            player_paginated = paginator.page(page)
        except PageNotAnInteger:
            player_paginated = paginator.page(1)
        except EmptyPage:
            player_paginated = paginator.page(paginator.num_pages)
        context['player'] = player_paginated
        return render(request, 'players/player_manage.html', context)
    else:
        try:
            if 'player_delete' in request.POST:
                player_id = request.POST.get('player_delete')
                player_delete = Player.objects.get(id=player_id)
                status = verificar_foto(str(player_delete.photo))
                if status:
                    player_delete.photo.delete()
                player_delete.delete()
                messages.success(request, "Atleta removido com sucesso!")
            return redirect('player_manage')
        except Exception as e: messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
        return redirect('player_manage')

@login_required(login_url="login")
@terms_accept_required
@permission_required('app.change_player', raise_exception=True)
def player_edit(request, id):
    try:
        campus = Campus_types.choices
        player = get_object_or_404(Player, id=id)
        if request.method == 'GET':
            return render(request, 'players/player_edit.html', {'player': player, 'campus': campus})            
        else:
            print(request.FILES)

            player.name = request.POST.get('name')
            player.sexo = request.POST.get('sexo')
            player.registration = request.POST.get('registration')
            player.cpf = request.POST.get('cpf')
            photo = request.FILES.get('photo')
            print("verifica:")
            if photo:
                print("photo")
                status = type_file(request, ['.png','.jpg','.jpeg'], photo, 'A photo anexada não é do tipo png, jpg ou jpeg, considere converte-la em um desses tipos.')
                if not status:
                    print("status")
                    status_photo = verificar_foto(str(player.photo))
                    if status_photo:
                        player.photo.delete()
                    player.photo = photo
            campus_id = request.POST.get('campus')
            if campus_id:
                player.campus = campus_id
            player.save()
            messages.success(request, "Os dados do atleta foram atualizados com sucesso!")
            return redirect('player_manage')
    except Exception as e: messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
    return redirect('Home')

@time_restriction()
@login_required(login_url="login")
@terms_accept_required
@permission_required('app.view_team_sport', raise_exception=True)
def team_manage(request):
    context = {}

    current_get_params = request.GET.urlencode()

    if request.method == "GET":
        if request.user.type == 0: 
            context['users'] = User.objects.all()
            context['events'] = Event.objects.all()   

        e = request.GET.get("e")
        t = request.GET.get("t")
        q = request.GET.get("q")

        if q and e and t:
            context['teams'] = Team.objects.filter(event__id=e)
            context['events_sport'] = Event_sport.objects.filter(event__id=e)
            context['team_sports'] = Team_sport.objects.filter(team__id=t)
            context['team'] = Team.objects.get(id=t)
            context['select_event'] = e
            context['event'] = Event.objects.get(id=e)

        elif e and t:
            context['teams'] = Team.objects.filter(event__id=e)
            context['events_sport'] = Event_sport.objects.filter(event__id=e)
            context['team_sports'] = Team_sport.objects.filter(team__id=t, team__event__id=e)
            context['team'] = Team.objects.get(id=t)
            context['voluntarys'] = Voluntary.objects.filter(event__id=e, type_voluntary=1)
            context['select_event'] = e
            context['event'] = Event.objects.get(id=e)

        elif e:
            context['teams'] = Team.objects.filter(event__id=e)
            context['events_sport'] = Event_sport.objects.filter(event__id=e)
            context['select_event'] = e
            context['event'] = Event.objects.get(id=e)

        elif t and request.user.type == 1:
            context['teams'] = Team.objects.filter(event__id=request.user.event_user.id)
            context['events_sport'] = Event_sport.objects.filter(event__id=request.user.event_user.id)
            context['team'] = Team.objects.get(id=t)
            context['team_sports'] = Team_sport.objects.filter(team__id=t)
            context['users'] = User.objects.filter(event_user=request.user.event_user)

        elif request.user.type == 1:
            print("eita")
            context['teams'] = Team.objects.filter(event__id=request.user.event_user.id)
            context['events_sport'] = Event_sport.objects.filter(event__id=request.user.event_user.id)

        elif request.user.type == 2:
            context['team'] = Team.objects.get(id=request.user.team.id)
            context['team_sports'] = Team_sport.objects.filter(team__id=request.user.team.id)
            context['events_sport'] = Event_sport.objects.filter(event__id=request.user.event_user.id)
        
        print("passou: ", context)
        return render(request, 'team/team_manage.html', context)

    else:
        print(request.POST)
        if 'add-team' in request.POST:
            name = request.POST.get("name")
            color = request.POST.get("color")
            description = request.POST.get("description")
            photo = request.FILES.get("photo")
            event = Event.objects.get(id=request.POST.get("add-team"))
            if not name or not photo or not event:
                messages.error(request,"Você precisa cadastrar todos os dados obrigatórios.")
            else:
                team = Team.objects.create(name=name, description=description, photo=photo, event=event)
                if color: team.color = str(color)
                team.save()
        elif 'add-team-sport' in request.POST:
            team = Team.objects.get(id=request.POST.get("add-team-sport"))
            sport = Event_sport.objects.get(id=request.POST.get("sport_adm_id"))
            sexo = int(request.POST.get("sexo_adm_id"))
            if not Team_sport.objects.filter(team=team, sport=sport, sexo=sexo, event=sport.event):
                if sexo == 0 and not sport.masc or sexo == 1 and not sport.fem or sexo == 2 and not sport.mist:
                    messages.error(request,"O esporte escolhido não está disponível para este sexo. Em caso de dúvidas, consulte o regulamento.")
                else:
                    team_sport = Team_sport.objects.create(team=team, sport=sport, sexo=sexo, event=sport.event)
                    if request.POST.get("technitian"):
                        team_sport.technitian = Voluntary.objects.get(id=request.POST.get("technitian"))
                    team_sport.save()
                    messages.success(request,"Esporte cadastrado com sucesso, adicione atletas.")
            else:
                messages.info(request,"O esporte já existe, adicione atletas.")
        elif 'edit-team' in request.POST:
            team = Team.objects.get(id=request.POST.get("edit-team"))
            team.name = request.POST.get("edit-name")
            team.color = request.POST.get("edit-color")
            if request.FILES.get("edit-logo"):
                team.photo = request.FILES.get("edit-logo")
            team.description = request.POST.get("edit-description")
            if request.POST.get('edit-status') == 'on': team.status = True
            else: team.status = False
            team.save()

        elif 'team-data' in request.POST:
            team_id = request.POST.get('team-data')
            team = Team.objects.get(id=team_id)
            cont = {
                'now': timezone.now(),
                'team': team,
                'user': request.user,
                'logo_morea': request.build_absolute_uri('/static/images/logo_atum.png')
            }

            if team.event:
                cont['logo_ifs'] = request.build_absolute_uri(team.event.logo.url)
            else:
                cont['logo_ifs'] = request.build_absolute_uri('/static/images/logo-jiifs-2025.jpg')
            logo_ifs = cont['logo_ifs']
            from reportlab.lib.utils import ImageReader


            img = ImageReader(logo_ifs)
            largura, altura = img.getSize()

            if largura == altura:
                print("Quadrada")
                cont['logo_event_type'] = 0
            elif largura > altura:
                print("Retangular horizontal")
                cont['logo_event_type'] = 1
            else:
                print("Retangular vertical")
                cont['logo_event_type'] = 2
            name_html = 'data-base-teams'
            name_pdf = f'relatório do {team.name}'
            
            teams = Team_sport.objects.filter(team=team).prefetch_related(Prefetch('players', queryset=Player_team_sport.objects.select_related('player'))).order_by('sport__sport', '-sexo')

            cont['teams'] = teams
            print(cont)
            html_string = render_to_string(f'generator/{name_html}.html', cont)

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="{name_pdf}.pdf"'
            # response['Content-Disposition'] = f'attachment; filename="{name_pdf}.pdf"'
            HTML(string=html_string).write_pdf(response)

            return response
        
        elif 'team_sport_delete' in request.POST:
            team_sport_id = request.POST.get('team_sport_delete')
            team_sport_delete = Team_sport.objects.get(id=team_sport_id)
            players_team_sport = Player_team_sport.objects.filter(team_sport=team_sport_delete)
            print(players_team_sport)
            if players_team_sport:
                for i in players_team_sport:
                    i.delete()     
                    print("apagado player somente do")
                    if not Player_team_sport.objects.filter(player=i.player).exists():
                        
                        print("APAGANDO JOGADOR: ", i.player.name)
                        status = verificar_foto(str(i.player.photo))
                        if status:
                            i.player.photo.delete()
                        i.player.bulletin.delete()
                        i.player.rg.delete()
                        i.player.delete()     
            team_sport_delete.delete()
            if not Team_sport.objects.filter(team=team_sport_delete.team.id):
                pass
                #Team.objects.get(id=team_sport_delete.team.id).delete()

        if current_get_params:
            return redirect(f"{reverse('team_manage')}?{current_get_params}")
        else:
            return redirect('team_manage')

@login_required(login_url="login") 
def theme_manage(request):
    return render(request, 'settings/theme.html')

@login_required(login_url="login")
@terms_accept_required
@permission_required('app.change_team_sport', raise_exception=True)
def team_edit(request, id):
    team_sport = get_object_or_404(Team_sport, id=id)
    sport = Sport_types.choices
    campus = Campus_types.choices
    sexo = Sexo_types.choices
    users = User.objects.all()
    if request.method == 'GET': 
        return render(request, 'team/team_edit.html', {'team_sport': team_sport, 'campus': campus, 'sport': sport, 'sexo': sexo, 'users': users})
    else:
        try:
            team_sport.sport = request.POST.get('sport')
            team_sport.sexo = request.POST.get('sexo')
            team_sport.admin = User.objects.get(id=request.POST.get('user'))
            team_sport.team.campus = request.POST.get('campus')
            for i in campus: 
                if i[0] == int(request.POST.get('campus')): team_sport.team.name = i[1]
            team_sport.team.save() 
            team_sport.save() 
            messages.success(request, 'Alteração feita com sucesso!')
            return redirect('team_manage') 
        except (TypeError, ValueError): messages.error(request, 'Um valor foi informado incorretamente!')
        except IntegrityError as e: messages.error(request, 'Algumas informações não foram preenchidas :(')
        except Exception as e: messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
    return redirect('team_manage') 

def calcular_idade(data_nasc):
    hoje = date.today()
    return hoje.year - data_nasc.year - (
        (hoje.month, hoje.day) < (data_nasc.month, data_nasc.day)
    )

@time_restriction()
@login_required(login_url="login")
@terms_accept_required
@permission_required('app.view_player_team_sport', raise_exception=True)
def team_players_manage(request, id):
    team_sport = get_object_or_404(Team_sport, id=id)

    # atualiza status de inscrição
    count_players = Player_team_sport.objects.filter(team_sport=team_sport).count()
    if count_players >= team_sport.sport.min_sport and not team_sport.status:
        team_sport.status = True
        team_sport.save()
    elif count_players < team_sport.sport.min_sport and team_sport.status:
        team_sport.status = False
        team_sport.save()

    user = User.objects.get(id=request.user.id)

    if request.method == "GET":
        player_team_sport = (
            Player_team_sport.objects
            .select_related('player', 'team_sport')
            .filter(team_sport=id)
        )
        context = {
            'player_team_sport': player_team_sport,
            'sexos': Sexo_types.choices,
            'team_sport': team_sport,
            'events': Event.objects.all(),
            'events_unit': Event_unit.objects.filter(event=team_sport.event),
        }
        return render(request, 'team/team_players_manage.html', context)

    # ── POST ──────────────────────────────────────────────────────────────────

    # ── remover jogador ───────────────────────────────────────────────────────
    if request.POST.get("player_delete"):
        player_id = request.POST.get('player_delete')
        Player_team_sport.objects.filter(team_sport=id, player=player_id).delete()
        if not Player_team_sport.objects.filter(player=player_id).exists():
            Player.objects.filter(id=player_id).delete()
        count_players = Player_team_sport.objects.filter(team_sport=team_sport).count()
        if count_players < team_sport.sport.min_sport and team_sport.status:
            team_sport.status = False
            team_sport.save()
        return redirect('team_players_manage', team_sport.id)

    # ── editar jogador existente ───────────────────────────────────────────────
    elif request.POST.get("edit-name"):
        player = Player.objects.get(id=request.POST.get("edit-player-id"))
        player.name = request.POST.get("edit-name")

        if team_sport.team.event.general_need_unit:
            player.unit = Event_unit.objects.get(id=int(request.POST.get('edit-unit')))

        if team_sport.team.event.player_need_address:
            player.address = request.POST.get('edit-address')

        if team_sport.team.event.player_need_registration:
            player.registration = request.POST.get("edit-registration")
            
        if team_sport.team.event.player_need_date_nasc:
            date_nasc = datetime.strptime(request.POST.get("edit-date"), "%Y-%m-%d").date()
            idade = calcular_idade(date_nasc)

            if idade < team_sport.team.event.age:
                messages.error(request, "Não foi possível atualizar: idade abaixo do mínimo.")
                return redirect('team_players_manage', team_sport.id)

            if idade > team_sport.team.event.age_max:
                messages.error(request, "Não foi possível atualizar: idade acima do máximo permitido.")
                return redirect('team_players_manage', team_sport.id)

            player.date_nasc = date_nasc

        if team_sport.team.event.player_need_photo and request.FILES.get("edit-photo"):
            player.photo = request.FILES.get("edit-photo")

        if team_sport.team.event.player_need_course and request.POST.get("edit-course"):
            player.course = request.POST.get('edit-course')

        if team_sport.team.event.player_need_cpf:
            player.cpf = (request.POST.get('edit-cpf') or '').replace("-", "").replace(".", "")

        if team_sport.team.event.player_need_cep:
            player.cep = (request.POST.get('edit-cep') or '').replace("-", "").replace(".", "").replace(" ", "")

        if team_sport.team.event.player_need_municipality:
            player.municipality = request.POST.get('edit-municipality')

        if team_sport.team.event.player_need_photo_goal and request.FILES.get("edit-photo-goal"):
            player.photo_goal = request.FILES.get("edit-photo-goal")

        if team_sport.team.event.player_need_bulletin and request.FILES.get("edit-bulletin"):
            player.bulletin = request.FILES.get("edit-bulletin")

        if team_sport.team.event.player_need_rg and request.FILES.get("edit-rg"):
            player.rg = request.FILES.get("edit-rg")

        player.save()
        messages.success(request, "Atleta atualizado com sucesso!")
        return redirect('team_players_manage', team_sport.id)

    # ── buscar jogador existente no sistema ────────────────────────────────────
    elif 'search' in request.POST:
        count_players = Player_team_sport.objects.filter(team_sport=team_sport).count()
        if count_players >= team_sport.sport.max_sport:
            messages.error(request, "O time atingiu o limite de atletas nessa modalidade!")
            return redirect('team_players_manage', team_sport.id)

        qe = request.POST.get('search')
        admin_filter = None if user.type in (0, 1) else user

        base_qs = _player_queryset_for_team(
            team_sport, user, team_sport.team.event, admin_user=admin_filter
        )

        if qe.isdigit():
            player_filter = base_qs.filter(registration=int(qe))
        else:
            player_filter = base_qs.filter(name__icontains=qe)

        if player_filter.count() > 1:
            messages.error(
                request,
                f"{player_filter.count()} atletas encontrados — seja mais preciso "
                f"(use nome completo ou matrícula)."
            )
        elif not player_filter.exists():
            messages.error(request, "Atleta não encontrado ou incompatível com o sexo do time.")
        else:
            player = player_filter.first()

            # verificação final de compatibilidade
            ok, err = _check_gender_compatibility(player, team_sport, user)
            if not ok:
                messages.error(request, err)
            elif Player_team_sport.objects.filter(player=player, team_sport=team_sport).exists():
                messages.info(request, f"O atleta '{player.name}' já está nessa modalidade.")
            else:
                Player_team_sport.objects.create(player=player, team_sport=team_sport)
                messages.success(request, f"Atleta '{player.name}' adicionado com sucesso!")

        return redirect('team_players_manage', team_sport.id)

    # ── criar novo jogador ─────────────────────────────────────────────────────
    elif 'name' in request.POST:
        count_players = Player_team_sport.objects.filter(team_sport=team_sport).count()
        if count_players >= team_sport.sport.max_sport:
            messages.error(request, "O time atingiu o limite de atletas nessa modalidade!")
            return redirect('team_players_manage', team_sport.id)

        # validações de arquivo
        if team_sport.team.event.player_need_date_nasc:
            date_nasc = datetime.strptime(request.POST.get('date'), "%Y-%m-%d").date()
            idade = calcular_idade(date_nasc)

            if idade < team_sport.team.event.age:
                messages.error(request, "O atleta não pode ser cadastrado: idade abaixo do mínimo.")
                return redirect('team_players_manage', team_sport.id)

            if idade > team_sport.team.event.age_max:
                messages.error(request, "O atleta não pode ser cadastrado: idade acima do máximo permitido.")
                return redirect('team_players_manage', team_sport.id)

        for field, exts, label in [
            ('photo',      ['.png', '.jpg', '.jpeg'], 'A foto deve ser PNG, JPG ou JPEG.'),
            ('photo_goal', ['.png', '.jpg', '.jpeg'], 'A foto gol deve ser PNG, JPG ou JPEG.'),
            ('bulletin',   ['.pdf'],                  'O boletim deve ser PDF.'),
            ('rg',         ['.png', '.jpg', '.jpeg', '.pdf', '.docx'], 'Formato de RG inválido.'),
        ]:
            if team_sport.team.event.__dict__.get(f'player_need_{field}', False):
                f = request.FILES.get(field)
                if f and type_file(request, exts, f, label):
                    return redirect('team_players_manage', team_sport.id)

        # sexo do jogador
        sexo_jogador = None

        if team_sport.sexo in (0, 1):
            # time masculino ou feminino: herda automaticamente
            sexo_jogador = team_sport.sexo
        else:
            # time misto: sexo é obrigatório
            raw = request.POST.get('sexo')
            if raw in (None, '', 'None'):
                messages.error(request, "Em times mistos, selecione o sexo do atleta.")
                return redirect('team_players_manage', team_sport.id)

            try:
                sexo_jogador = int(raw)
            except (TypeError, ValueError):
                messages.error(request, "Selecione um sexo válido para o atleta.")
                return redirect('team_players_manage', team_sport.id)

            if sexo_jogador not in (0, 1):
                messages.error(request, "Selecione um sexo válido para o atleta.")
                return redirect('team_players_manage', team_sport.id)

        # validação de gênero antes de criar
        if not user.is_superuser:
            if team_sport.sexo in (0, 1) and sexo_jogador != team_sport.sexo:
                messages.error(
                    request,
                    f"O sexo do atleta é incompatível com o time "
                    f"({SEXO_NAMES[team_sport.sexo]})."
                )
                return redirect('team_players_manage', team_sport.id)

        name = request.POST.get('name')
        photo      = request.FILES.get('photo')
        photo_goal = request.FILES.get('photo_goal')
        bulletin   = request.FILES.get('bulletin')
        rg         = request.FILES.get('rg')

        player, created = Player.objects.get_or_create(
            name=name, admin=user, event=team_sport.event,
        )

        if team_sport.team.event.general_need_unit:
            if user.type == 2:
                player.unit = Event_unit.objects.get(id=user.unit.id)
            else:
                player.unit = Event_unit.objects.get(id=int(request.POST.get('unit')))

        if team_sport.team.event.player_need_date_nasc:
            player.date_nasc = date_nasc
        if team_sport.team.event.player_need_address:
            player.address = request.POST.get('address')
        if team_sport.team.event.player_need_registration:
            player.registration = request.POST.get('registration')
        if team_sport.team.event.player_need_cpf:
            player.cpf = (request.POST.get('cpf') or '').replace("-", "").replace(".", "")
        if team_sport.team.event.player_need_photo and photo:
            player.photo = photo
        if team_sport.team.event.player_need_course:
            player.course = request.POST.get('course')
        if team_sport.team.event.player_need_cep:
            player.cep = (request.POST.get('cep') or '').replace("-", "").replace(".", "").replace(" ", "")
        if team_sport.team.event.player_need_municipality:
            player.municipality = request.POST.get('municipality')
        if team_sport.team.event.player_need_photo_goal and photo_goal:
            player.photo_goal = photo_goal
        if team_sport.team.event.player_need_bulletin and bulletin:
            player.bulletin = bulletin
        if team_sport.team.event.player_need_rg and rg:
            player.rg = rg

        player.sexo = sexo_jogador

        if player.sexo is None:
            messages.error(request, "Não foi possível cadastrar o atleta sem sexo definido.")
            return redirect('team_players_manage', team_sport.id)

        player.save()

        _, vínculo_criado = Player_team_sport.objects.get_or_create(
            player=player, team_sport=team_sport
        )
        if vínculo_criado:
            messages.success(request, "Atleta cadastrado com sucesso!")
        else:
            messages.info(request, "O atleta já estava nessa modalidade.")

        # reavalia status
        count_players = Player_team_sport.objects.filter(team_sport=team_sport).count()
        if count_players >= team_sport.sport.min_sport and not team_sport.status:
            team_sport.status = True
            team_sport.save()

        return redirect('team_players_manage', team_sport.id)

    return redirect('team_players_manage', team_sport.id)

@login_required(login_url="login")
@terms_accept_required
@permission_required('app.add_player_team_sport', raise_exception=True)
def add_player_team(request, id):
    team = get_object_or_404(Team_sport, id=id)
    players = Player.objects.all()
    if request.method == 'GET':
        if not players: messages.info(request, "Não tem nenhum atleta cadastrado no sistema!")
        return render(request, 'add_players_team.html', {'players': players,'team': team}) 
    else:
        try:
            player = request.POST.getlist('input-checkbox')
            for i in player:
                player = Player.objects.get(id=i)
                Player_team_sport.objects.create(player=player, team_sport=team)
        except (TypeError, ValueError):
            messages.error(request, 'Um valor foi informado incorretamente!')
        except IntegrityError as e:
            messages.error(request, 'Algumas informações não foram preenchidas :(')
        except Exception as e:
            messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
        return redirect('team_players_manage', team.id)
    
@login_required(login_url="login")
@terms_accept_required
@permission_required('app.view_match', raise_exception=True)
def matches_manage(request):
    try:
        matchs = Match.objects.all().prefetch_related('teams__team')
        sport = Sport_types.choices
        context = [
            {
                'match': match,
                'sport':sport,
                'times': list(match.teams.all()),
                
            }
            for match in matchs
        ]
        if request.method == "GET":
            if not context:
                print("Não há nenhuma partida cadastrada!")
            return render(request, 'matches/matches_manage.html',{'context': context})
        else:
            match_id = request.POST.get('match_delete')
            match_delete = Match.objects.get(id=match_id)
            if match_delete.sport == 1:
                if Volley_match.objects.filter(id=match_delete.volley_match.id):
                    volley_match = Volley_match.objects.get(id=match_delete.volley_match.id)
                    matches = Match.objects.filter(volley_match=volley_match.id)
                    if len(matches) < 2:
                        volley_match.delete()
            match_delete.delete()
            return redirect('matches_manage')
    except Exception as e:
        messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
        return redirect('matches_manage')

@login_required(login_url="login")
@terms_accept_required
@permission_required('app.change_match', raise_exception=True)
def matches_edit(request, id):
    try:
        match = get_object_or_404(Match, id=id)
        team_matchs = Team_match.objects.filter(match=match)
        team = Team.objects.all()
        team_match_a = team_matchs[0]
        team_match_b = team_matchs[1]
        match = get_object_or_404(Match, id=id)
        sport = Sport_types.choices
        if match.sexo != 1 or 0:
            match_disable = True
        else:
            match_disable = False
        
        context = {
            'match': match, 
            'sport': sport,
            'team': team,
            'team_match_a': team_match_a,
            'team_match_b': team_match_b,
            'match_disable': match_disable,
        }
        if request.method == "GET":
            return render(request, 'matches/matches_edit.html', context)
        else:
            if 'excluir' in request.POST:
                print("certin")
                match.delete()
                if match.sport == 1:
                    volley_match = Volley_match.objects.get(id=match.volley_match.id)
                    volley_match.delete()
                team_match_a.delete()
                team_match_b.delete()
                return redirect('matches_manage')
            else:
                sport_select = int(request.POST.get('sport'))
                match.sport = sport_select
                match.sport = sport
                match.sexo = request.POST.get('sexo')
                match.time_match = request.POST.get('datatime')
                team_a = request.POST.get('team_a')
                team_b = request.POST.get('team_b')
                team_match_a.team = get_object_or_404(Team, id=team_a)
                team_match_b.team = get_object_or_404(Team, id=team_b)
                team_match_a.save()
                team_match_b.save()
                match.time_match = request.POST.get('datetime')
                match.save()
    except (TypeError, ValueError):
        messages.error(request, 'Um valor foi informado incorretamente!')
    except IntegrityError as e:
        messages.error(request, 'Algumas informações não foram preenchidas :(')
    except Team.DoesNotExist:
        messages.error(request, 'Um dos times não foi informado ou é inexistente!')
    except Exception as e:
        messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
    return redirect('matches_manage')

@login_required(login_url="login")
@terms_accept_required
def games(request):
    if request.method == "GET":
        context = {
            'team': Team.objects.all(),
            'sport': Sport_types.choices,
            'events': Event.objects.all(),
            'phase_types': Phase_types.choices,
            'sexo': Sexo_types.choices,
        }

        selected_event = None

        if 'e' in request.GET and request.GET['e'] != '':
            selected_event = Event.objects.get(id=request.GET['e'])
            context['select_event'] = request.GET['e']
            context['phases'] = Phase.objects.filter(event__event__id=request.GET['e']).order_by('name','event','sexo')
            context['groups'] = Group_phase.objects.filter(phase__event__event__id=request.GET['e']).order_by('phase__name','phase__event','phase__sexo')
            context['event_sports'] = Event_sport.objects.filter(event=selected_event)
            context['teams'] = Team.objects.filter(event=selected_event)
        elif request.user.event_user:
            selected_event = request.user.event_user
            context['event_sports'] = Event_sport.objects.filter(event=request.user.event_user)
            context['phases'] = Phase.objects.filter(event__event=request.user.event_user).order_by('name','event','sexo')
            context['groups'] = Group_phase.objects.filter(phase__event__event=request.user.event_user).order_by('phase__name','phase__event','phase__sexo')

        if selected_event:
            matches = Match.objects.filter(event__id=selected_event.id).prefetch_related('teams__team').order_by('time_match')
        else:
            context['phases'] = []
            context['groups'] = []
            matches = Match.objects.all().prefetch_related('teams__team').order_by('time_match')

        context['context'] = [
            {
                'match': match,
                'times': list(match.teams.all()),
                'points_a': Point.objects.filter(team_match=match.teams.first()).count(),
                'points_b': Point.objects.filter(team_match=match.teams.last()).count(),
            }
            for match in matches
        ]

        return render(request, 'games.html', context)

    elif 'change_match' in request.POST:
        match = Match.objects.get(id=request.POST.get('change_match'))

        if Match.objects.filter(status=1, event=match.event).exists():
            messages.info(request, "Já existe uma partida em andamento. Finalize-a antes de iniciar outra.")
        elif match.status == 0:
            if match.volley_match:
                volley_match = Volley_match.objects.get(id=match.volley_match.id)
                volley_match.status = 1
                volley_match.save()
            match.status = 1
            match.save()
            return redirect('scoreboard', match.event.id)
        else:
            messages.info(request, "A partida já foi finalizada.")
        return redirect('games')

    # 2️⃣ Criar nova FASE
    elif 'create_phase' in request.POST:
        if not request.user.has_perm('app.add_phase'):
            messages.error(request, "Você não tem permissão para criar fases.")
            return redirect('games')
        print(request.POST)
        event_id = int(request.POST.get('event_sport'))
        name = int(request.POST.get('name'))
        sexo = int(request.POST.get('sexo_phase'))
        if not event_id or not name or not sexo:
            if not name == 0 and not name or not sexo == 0 and not sexo:
                messages.error(request, "Dados insuficientes para criar a fase.")
                return redirect('games')

        event_sport = Event_sport.objects.get(id=event_id)
        print("event_S", event_sport)
        if not event_sport:
            messages.error(request, "Evento ou esporte não encontrado.")
            return redirect('games')

        Phase.objects.create(event=event_sport, name=name, sexo=sexo)
        messages.success(request, "Fase criada com sucesso!")
        return redirect(f"{reverse('games')}?e={event_sport.event.id}")

    # 3️⃣ Criar novo GRUPO
    elif 'create_group' in request.POST:
        if not request.user.has_perm('app.add_group_phase'):
            messages.error(request, "Você não tem permissão para criar grupos.")
            return redirect('games')

        phase_id = request.POST.get('phase')
        group_name = request.POST.get('group_name')

        if not phase_id:
            messages.error(request, "Preencha todos os campos para criar o grupo.")
            return redirect('games')

        phase = Phase.objects.get(id=phase_id)
        Group_phase.objects.create(phase=phase, name=group_name)
        messages.success(request, "Grupo criado com sucesso!")
        return redirect(f"{reverse('games')}?e={phase.event.event.id}")

    elif 'time_a' in request.POST and 'time_b' in request.POST:

        # 4️⃣ Criar nova PARTIDA
        event_sport = Event_sport.objects.get(id=int(request.POST.get('sport')))
        sport_id = event_sport.sport
        sexo = request.POST.get('sexo')
        team_a_id = request.POST.get('time_a')
        team_b_id = request.POST.get('time_b')
        datetime = request.POST.get('datetime')
        group_phase_id = request.POST.get('group')
        location = request.POST.get('location')

        if group_phase_id:
            if sport_id != Group_phase.objects.get(id=group_phase_id).phase.event.sport:
                messages.error(request, "O grupo precisa corresponder ao esporte.")
                return redirect('games')

        # Define o evento
        if not request.user.event_user:
            if 'e' in request.GET and request.GET['e'] != '':
                event = Event.objects.get(id=request.GET['e'])
            else:
                messages.error(request, "Selecione um evento válido.")
                return redirect('games')
        else:
            event = request.user.event_user

        # Validações
        if team_a_id == team_b_id:
            messages.error(request, "Você não pode criar uma partida com times iguais!")
            return redirect('games')

        team_a = Team.objects.get(id=team_a_id)
        team_b = Team.objects.get(id=team_b_id)

        team_sport_a = Team_sport.objects.filter(team=team_a, sport=event_sport, sexo=sexo).first()
        team_sport_b = Team_sport.objects.filter(team=team_b, sport=event_sport, sexo=sexo).first()

        if not team_sport_a or not team_sport_b:
            messages.error(request, "Algum time não está cadastrado na modalidade selecionada!")
            return redirect('games')
        
        if sport_id in [1, 2]:
            volley_match = Volley_match.objects.create(status=0, event=event)
            volley_match.save()
            match, created = Match.objects.get_or_create(
                sport=sport_id,
                sexo=sexo,
                time_match=datetime,
                volley_match=volley_match,
                event=event,
                defaults={
                    'group_phase_id': group_phase_id or None,
                    'location': location or "",
                }
            )
        else:
            match, created = Match.objects.get_or_create(
                sport=sport_id,
                sexo=sexo,
                time_match=datetime,
                event=event,
                defaults={
                    'group_phase_id': group_phase_id or None,
                    'location': location or "",
                }
            )

        if created:
            Team_match.objects.create(match=match, team=team_a)
            Team_match.objects.create(match=match, team=team_b)
            messages.success(request, "Partida cadastrada com sucesso!")
        else:
            messages.info(request, f"Essa partida já foi cadastrada! Identificação: #{match.id}")

        team_matches = Team_match.objects.filter(match=match)
        team_match_a = team_matches[0]
        team_match_b = team_matches[1]
  
        players_match_a = Player_match.objects.filter(team_match=team_match_a)
        players_match_b = Player_match.objects.filter(team_match=team_match_b)


        player_team_sport_a = Player_team_sport.objects.filter(team_sport=team_sport_a)
        player_team_sport_b = Player_team_sport.objects.filter(team_sport=team_sport_b)

        for i in player_team_sport_a:
            Player_match.objects.get_or_create(player=i.player, match=match, team_match=team_match_a)
        for i in player_team_sport_b:
            Player_match.objects.get_or_create(player=i.player, match=match, team_match=team_match_b)

        for i in players_match_a:
            if not Player_team_sport.objects.filter(player=i.player, team_sport=team_sport_a).exists():
                i.delete()
        for i in players_match_b:
            if not Player_team_sport.objects.filter(player=i.player, team_sport=team_sport_b).exists():
                i.delete()

    elif 'sumula' in request.POST:
        print(request.POST)
        match = Match.objects.get(id=request.POST.get('sumula'))
        match_referee = Match_referee.objects.filter(match=match)
        team_match = Team_match.objects.filter(match=match)
        players_match_a = Player_match.objects.filter(team_match=team_match[0])
        players_match_b = Player_match.objects.filter(team_match=team_match[1])
        team_sport_a = Team_sport.objects.get(team=team_match[0].team, sport__sport=team_match[0].match.sport, sexo=team_match[0].match.sexo)
        team_sport_b = Team_sport.objects.get(team=team_match[1].team, sport__sport=team_match[1].match.sport, sexo=team_match[1].match.sexo)
        point_a = Point.objects.filter(team_match=team_match[0])
        point_b = Point.objects.filter(team_match=team_match[1])
        replacements = Replacement.objects.filter(team_match__match=match)
        players_a = [
            {
            'name': i.player.name,
            'number': i.player_number,
            'card_r': Penalties.objects.filter(player=i.player, team_match=team_match[0], type_penalties=0).count(),
            'card_y': Penalties.objects.filter(player=i.player, team_match=team_match[0], type_penalties=1).count(),
            'point': Point.objects.filter(player=i.player, team_match=team_match[0]).count(),
            }
            for i in players_match_a
        ]
        players_b = [
            {
            'name': i.player.name,
            'number': i.player_number,
            'card_r': Penalties.objects.filter(player=i.player, team_match=team_match[0], type_penalties=0).count(),
            'card_y': Penalties.objects.filter(player=i.player, team_match=team_match[0], type_penalties=1).count(),
            'point': Point.objects.filter(player=i.player, team_match=team_match[0]).count(),
            }
            for i in players_match_b
        ]

        authenticity = generate_authenticity(f"Súmula gerada por {request.user.username} da partida entre {team_match[0].team.name} e {team_match[1].team.name}", match.event)
        link = f"https://{request.get_host()}/autenticar?code={authenticity.code}"
        qr = qrcode.make(link)
        print(link)
        buffer = BytesIO()
        qr.save(buffer, format='PNG')
        buffer.seek(0)

        img_base64 = base64.b64encode(buffer.getvalue()).decode()
        context = {
            'match':match,
            'team_match_a':team_match[0],
            'team_match_b':team_match[1],
            'players_a':players_a,
            'players_b':players_b,
            'team_sport_a':team_sport_a,
            'team_sport_b':team_sport_b,
            'point_a':point_a,
            'point_b':point_b,
            'replacements': replacements,
            'user': request.user,
            'match_referee': match_referee,
            'authenticity': authenticity,
            'qr_code': img_base64,
            'logo_ifs': request.build_absolute_uri('/static/images/logo-ifs-black.svg'),
            'logo_morea': request.build_absolute_uri('/static/images/logo-morea.svg'),
            
        }
        print(context)
        html_string = render_to_string('generator/sumula.html', context)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="sumula-{authenticity.number}.pdf"'
        #response['Content-Disposition'] = f'attachment; filename="sumula.pdf"'

        HTML(string=html_string).write_pdf(response)

        return response
    return redirect('games')

def generate_authenticity(name, event):
    list = []
    possibilidade = ["0","1","2","3","4","5","6","7","8","9","A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
    for i in range(4):
        number = random.randint(1, 8)
        chars = ''.join(random.sample(possibilidade, number))
        list.append(chars)
        if i == 3: list.append(f"-{number}")
        else: list.append(f"-{number}-")
    result = str(''.join(list))
    number = str(random.randint(1111111,9999999))
    authenticity = Authenticity.objects.create(name=name, event=event, code=result, number=number)
    authenticity.save()
    return authenticity

def authenticate_file(request):
    context = {}
    if 'code' in request.GET and request.GET.get('code') != '':
        code = str(request.GET.get('code'))
        if Authenticity.objects.filter(code=code):
            context['authenticity'] = Authenticity.objects.filter(code=code)[0]
        context['status'] = True
        print("code")
        
    return render(request, 'public/authenticate.html', context)

@login_required(login_url="login")
def get_teams(request):
    print("chegou")
    sport_id = request.GET.get('sport')
    sexo = request.GET.get('sexo')
    print("acordaaa", sport_id, sexo)
    # Exemplo: filtrando os times pelo esporte e sexo
    teams = Team_sport.objects.filter(sport__id=sport_id, sexo=sexo)
    data = {"teams": [{"id": t.team.id, "name": t.team.name} for t in teams]}
    return JsonResponse(data)

@login_required(login_url="login")
def get_groups(request):
    sport_id = request.GET.get('sport')
    groups = Group_phase.objects.filter(phase__event__id=sport_id).order_by('phase__name','phase__event','phase__sexo')
    data = [
        {
            "id": g.id,
            "name": g.name,
            "phase_name": g.phase.get_name_display(),
            "sexo": g.phase.get_sexo_display(),
        }
        for g in groups
    ]
    return JsonResponse({"groups": data})

@login_required(login_url="login")
def get_sexos(request):
    sport_id = request.GET.get('sport')
    esport = Event_sport.objects.get(id=sport_id)

    sexos = []
    if esport.masc:
        sexos.append({"value": 0, "label": "Masculino"})
    if esport.fem:
        sexos.append({"value": 1, "label": "Feminino"})
    if esport.mist:
        sexos.append({"value": 2, "label": "Misto"})

    return JsonResponse({"sexos": sexos})

@login_required(login_url="login")
@permission_required('app.view_match', raise_exception=True)
def match_settings(request, id_sport, id_match):
    match = Match.objects.get(id=id_match)
    current_get_params = request.GET.urlencode()
    if request.method == "GET":
        match=match
        time_pauses = Time_pause.objects.filter(match=match)
        assistance = Assistance.objects.filter(assis_to__team_match__match=match)

        team_match_a = Team_match.objects.filter(match=match)[0]
        team_match_b = Team_match.objects.filter(match=match)[1]
        player_a = Player_match.objects.filter(team_match=team_match_a)
        player_b = Player_match.objects.filter(team_match=team_match_b)
        points_a = Point.objects.filter(team_match=team_match_a)
        points_b = Point.objects.filter(team_match=team_match_b)
        penalties_a = Penalties.objects.filter(team_match=team_match_a)
        penalties_b = Penalties.objects.filter(team_match=team_match_b)
        match_referee = Match_referee.objects.filter(match=match)
        
        group_phases = Group_phase.objects.filter(phase__event__event=match.event)
        player_match = Player_match.objects.filter(match=match)
        team_match = Team_match.objects.filter(match=match)
        sports = Event_sport.objects.filter(event=match.event)
        status = Status.choices
        detailed = Detailed.choices
        sexos = Sexo_types.choices

        context = {
            'match': match,
            'player_a': player_a,
            'player_b': player_b,
            'points_a': points_a,
            'points_b': points_b,
            'penalties_a': penalties_a,
            'penalties_b': penalties_b,
            'team_match_a': team_match_a,
            'team_match_b': team_match_b,
            'time_pauses': time_pauses,
            'assistance': assistance,
            'match_referee': match_referee,

            'players_match': player_match,
            'teams_match': team_match,
            'group_phases': group_phases,
            'sports': sports,
            'status': status,
            'detailed': detailed,
            'sexos': sexos,
        }
        if match.sport in [1, 2]: 
            context['volley_matchs'] = Volley_match.objects.filter(event=match.event)

        return render(request, 'match_settings.html', context)
    else:
        if 'pauses_delete' in request.POST:
            pause = Time_pause.objects.get(id=request.POST.get('pauses_delete'))
            pause.delete()
        elif 'penalties_delete' in request.POST:
            penalties = Penalties.objects.get(id=request.POST.get('penalties_delete'))
            penalties.delete()
        elif 'point_delete' in request.POST:
            point = Point.objects.get(id=request.POST.get('point_delete'))
            point.delete()
        elif 'assistance_delete' in request.POST:
            assistance = Assistance.objects.get(id=request.POST.get('assistance_delete'))
            assistance.delete()
        elif 'referee_delete' in request.POST:
            referee = Match_referee.objects.get(id=request.POST.get('referee_delete'))
            referee.delete()
        elif 'sumula' in request.POST:
            match_referee = Match_referee.objects.filter(match=match)
            team_match = Team_match.objects.filter(match=match)
            players_match_a = Player_match.objects.filter(team_match=team_match[0])
            players_match_b = Player_match.objects.filter(team_match=team_match[1])
            team_sport_a = Team_sport.objects.get(team=team_match[0].team, sport__sport=team_match[0].match.sport, sexo=team_match[0].match.sexo)
            team_sport_b = Team_sport.objects.get(team=team_match[1].team, sport__sport=team_match[1].match.sport, sexo=team_match[1].match.sexo)
            point_a = Point.objects.filter(team_match=team_match[0])
            point_b = Point.objects.filter(team_match=team_match[1])
            replacements = Replacement.objects.filter(team_match__match=match)
            players_a = [
                {
                'name': i.player.name,
                'number': i.player_number,
                'card_r': Penalties.objects.filter(player=i.player, team_match=team_match[0], type_penalties=0).count(),
                'card_y': Penalties.objects.filter(player=i.player, team_match=team_match[0], type_penalties=1).count(),
                'point': Point.objects.filter(player=i.player, team_match=team_match[0]).count(),
                }
                for i in players_match_a
            ]
            players_b = [
                {
                'name': i.player.name,
                'number': i.player_number,
                'card_r': Penalties.objects.filter(player=i.player, team_match=team_match[0], type_penalties=0).count(),
                'card_y': Penalties.objects.filter(player=i.player, team_match=team_match[0], type_penalties=1).count(),
                'point': Point.objects.filter(player=i.player, team_match=team_match[0]).count(),
                }
                for i in players_match_b
            ]

            authenticity = generate_authenticity(f"Súmula gerada por {request.user.username} da partida entre {team_match[0].team.name} e {team_match[1].team.name}", match.event)
            link = f"https://{request.get_host()}/autenticar?code={authenticity.code}"
            qr = qrcode.make(link)
            print(link)
            buffer = BytesIO()
            qr.save(buffer, format='PNG')
            buffer.seek(0)

            img_base64 = base64.b64encode(buffer.getvalue()).decode()
            context = {
                'match':match,
                'team_match_a':team_match[0],
                'team_match_b':team_match[1],
                'players_a':players_a,
                'players_b':players_b,
                'team_sport_a':team_sport_a,
                'team_sport_b':team_sport_b,
                'point_a':point_a,
                'point_b':point_b,
                'replacements': replacements,
                'user': request.user,
                'match_referee': match_referee,
                'authenticity': authenticity,
                'qr_code': img_base64,
                'logo_ifs': request.build_absolute_uri('/static/images/logo-ifs-black.svg'),
                'logo_morea': request.build_absolute_uri('/static/images/logo-morea.svg'),
                
            }
            print(context)
            html_string = render_to_string('generator/sumula.html', context)

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="sumula-{authenticity.number}.pdf"'
            #response['Content-Disposition'] = f'attachment; filename="sumula.pdf"'

            HTML(string=html_string).write_pdf(response)

            return response
        elif 'location' in request.POST or 'add' in request.POST or 'winner' in request.POST:
            print(request.POST)
            if request.POST.get('sport'): match.sport = int(request.POST.get('sport'))
            if request.POST.get('sexo'): match.sexo = int(request.POST.get('sexo'))
            if request.POST.get('status'): match.status = int(request.POST.get('status'))
            if request.POST.get('detailed'): match.detailed = int(request.POST.get('detailed'))
            if request.POST.get('mvp'): match.mvp_player_player = Player.objects.get(id=request.POST.get('mvp'))
            else: match.mvp_player_player = None
            if request.POST.get('winner'): match.Winner_team = Team.objects.get(id=request.POST.get('winner')) 
            else: match.Winner_team = None
            if request.POST.get('group_phase'): match.group_phase = Group_phase.objects.get(id=request.POST.get('group_phase'))
            if request.POST.get('volley_match'): match.volley_match = Volley_match.objects.get(id=request.POST.get('volley_match'))
            if request.POST.get('time_match'): match.time_match = request.POST.get('time_match')
            if request.POST.get('time_start'): match.time_start = request.POST.get('time_start')
            if request.POST.get('time_end'): match.time_end = request.POST.get('time_end')
            if request.POST.get('location'): match.location = request.POST.get('location')

            if request.POST.get('add'): match.add = request.POST.get('add')
            if request.POST.get('observations'): match.observations = request.POST.get('observations')
            match.save()
        if current_get_params:
            return redirect(f"{reverse('match_settings', args=[id_sport, id_match])}?{current_get_params}")
        else:
            return redirect('match_settings', match.sport, match.id)
    


@login_required(login_url="login")
def manage_session(request):
    current_get_params = request.GET.urlencode()
    if request.method == "GET":
        context = {'users': User.objects.all()}
        if 'e' in request.GET and request.GET.get('e') != '':
            print("Né: ",request.GET.get('e'))
            user_session = User.objects.get(id=request.GET.get('e'))
            context['user_session'] = user_session
            context['sessions'] = UserSession.objects.filter(user=user_session)
        else:
            context['sessions'] = UserSession.objects.filter(user=request.user).select_related("session")

        return render(request, "manage_sessions.html", context)
    else:
        session_key = request.POST.get("session_key")
        if session_key and session_key != request.session.session_key:
            try:
                Session.objects.get(session_key=session_key).delete()
                UserSession.objects.filter(session__session_key=session_key).delete()
            except Session.DoesNotExist:
                pass
        if current_get_params:
            return redirect(f"{reverse('manage_sessions')}?{current_get_params}")
        else:
            return redirect("manage_sessions")

@login_required(login_url="login")
@terms_accept_required
@permission_required('app.view_customuser', raise_exception=True)
def user_manage(request):
    context = {
        'events': Event.objects.all(), 
    }
    if not request.user.type in [0]:
        context['type_user'] = Users_types.choices[2:]
        context['team'] = Team.objects.filter(event__id=request.user.event_user.id)
        context['event_unit'] = Event_unit.objects.filter(event__id=request.user.event_user.id)
        context['users_all'] = User.objects.filter(event_user=request.user.event_user)
    elif 'e' in request.GET and request.GET['e'] != '':
        context['type_user'] = Users_types.choices
        context['users_all'] = User.objects.filter(event_user=Event.objects.get(id=request.GET['e']))
        context['select_event'] = request.GET['e']
        context['event_unit'] = Event_unit.objects.filter(event__id=request.GET['e'])
        context['team'] = Team.objects.filter(event__id=request.GET['e'])
    else:
        context['type_user'] = Users_types.choices
        context['users_all'] = User.objects.filter(event_user=None)
    
    if request.method == "GET":
        return render(request, 'settings/user_manage.html', context)
    else:
        print(request.POST, request.FILES)
        if 'user_id' in request.POST:
            user = get_object_or_404(User, id=request.POST.get('user_id'))
            print(user.username)
            print(user.password)
            if request.POST.get('name'): user.username = str(request.POST.get('name'))
            if request.POST.get('password'):
                senha = request.POST.get('password') 
                print("trocando: ", senha)
                user.set_password(senha)
            print("uaii")
            if int(request.POST.get('event')) != 0: user.event_user = Event.objects.get(id=request.POST.get('event'))
            else: user.event_user = None
            if request.POST.get('telephone'): user.telefone = request.POST.get('telephone')
            if request.POST.get('email'): user.email = str(request.POST.get('email'))
            if request.POST.get('active') == 'on': user.is_active = True
            else: user.is_active = False
            if request.POST.get('type'): 
                if int(request.POST.get('type')) == 2:
                    if request.POST.get('team'):
                        user.team = Team.objects.get(id=request.POST.get('team'))
                        user.type = request.POST.get('type')
                    else:
                        messages.info(request, f"O time/ou o tipo não foi alterado porque faltou informar o time.")
                else:
                    user.type = request.POST.get('type')
                    user.team = None
                
            if request.FILES.get('photo'):
                if user.photo: 
                    status = verificar_foto(str(user.photo))
                    if status:
                        user.photo.delete()
                user.photo = request.FILES.get('photo')
            user.save()
            print("Nova: ",user.password)
            messages.success(request, f"{user.username} do sistema atualizado com sucesso!")
        elif 'name' in request.POST:
            name = request.POST.get('name')
            if request.POST.get('event') and int(request.POST.get('event')) != 0: event = Event.objects.get(id=request.POST.get('event'))
            elif request.user.type == 1: event = request.user.event_user
            else: event = None
            type = request.POST.get('type')
            team = request.POST.get('team')
            email = request.POST.get('email')
            telephone = request.POST.get('telephone')
            password = request.POST.get('password')
            photo = request.FILES.get('photo')
            if int(type) == 2:
                if team: User.objects.create_user(username=name, password=password, photo=photo, type=type, team=Team.objects.get(id=team), email=email, telefone=telephone, event_user=event)
                elif request.user.team: User.objects.create_user(username=name, password=password, photo=photo, type=type, team=request.user.team, email=email, telefone=telephone, event_user=event)
                else: messages.error(request, "Você não informou o time associado ao usuário.")
            else:
                User.objects.create_user(username=name, password=password, photo=photo, type=type, email=email, telefone=telephone, event_user=event )
                messages.success(request, f"{name} cadastrado do sistema com sucesso!")
        elif 'new_password' in request.POST:
            user = User.objects.get(id=request.POST.get('new_password'))
            list = [] 
            possibilidade = ["0","1","2","3","4","5","6","7","8","9","A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
            list.append(''.join(random.sample(possibilidade, 8)))
            result = str(''.join(list))

            user.set_password(result)
            
            user.save()
            messages.success(request, f"Sucesso! Senha gerada com exito.")
            
            cont = {
                'now': timezone.now(),
                'user': request.user,
                'logo_atum': request.build_absolute_uri('/static/images/logo_atum.png')
            }
            
            if user.event_user:
                cont['logo_ifs'] = request.build_absolute_uri(user.event_user.logo.url)
                cont['event'] = user.event_user
            else:
                cont['logo_ifs'] = request.build_absolute_uri('/static/images/logo-jiifs-2025.jpg')
            logo_ifs = cont['logo_ifs']
            cont['logo_ifs_ofc'] = request.build_absolute_uri('/static/images/logo-ifs.svg')
            cont['logo_morea'] = request.build_absolute_uri('/static/images/logo-morea-sports.svg')
            link = f"https://{request.get_host()}"
            qr = qrcode.make(link)
            print(link)
            buffer = BytesIO()
            qr.save(buffer, format='PNG')
            buffer.seek(0)
            cont['site'] = link

            img_base64 = base64.b64encode(buffer.getvalue()).decode()

            img = ImageReader(logo_ifs)
            largura, altura = img.getSize()

            if largura == altura:
                print("Quadrada")
                cont['logo_event_type'] = 0
            elif largura > altura:
                print("Retangular horizontal")
                cont['logo_event_type'] = 1
            else:
                print("Retangular vertical")
                cont['logo_event_type'] = 2

            name_html = 'data-welcome'
            name_pdf = f'boas vindas, {user.username.lower()}'

            cont['name'] = user.username
            cont['qrcode'] = img_base64
            cont['password'] = result
        
            html_string = render_to_string(f'generator/{name_html}.html', cont)

            response = HttpResponse(content_type='application/pdf')
            #response['Content-Disposition'] = f'inline; filename="{name_pdf}.pdf"'  #gerar e deixar no navegador
            response['Content-Disposition'] = f'attachment; filename="{name_pdf}.pdf"' #gerar e baixar no navegador

            HTML(string=html_string).write_pdf(response)
            return response
    
        elif 'user_delete' in request.POST:
            user_id = request.POST.get('user_delete')
            user_delete = User.objects.get(id=user_id)
            user_delete.delete()
            messages.info(request, f"{user_delete.username} removido do sistema com sucesso!")
        return redirect('user_manage')

@time_restriction()
@login_required(login_url="login")
@terms_accept_required
@permission_required('app.view_voluntary', raise_exception=True)
def voluntary_manage(request):
    user = User.objects.get(id=request.user.id)
    if user.is_staff:
        types = Type_service.choices
    else:
        types = [choice for choice in Type_service.choices if choice[0] != 4]
    if request.method == "GET":
        context = {
            'types': types,
            'users': User.objects.all(),
        }
        if not request.user.event_user: context['events'] = Event.objects.all()
        if 'e' in request.GET and request.GET['e'] != '' and 'q' in request.GET:
            context['voluntarys'] = Voluntary.objects.filter(name__icontains=request.GET['q'], event__id=request.GET['e'])
            context['events_unit'] = Event_unit.objects.filter(event__id=request.GET['e'])
        elif 'e' in request.GET and request.GET['e'] != '':
            context['voluntarys'] = Voluntary.objects.filter(event__id=request.GET['e'])
            context['select_event'] = request.GET['e']
            context['events_unit'] = Event_unit.objects.filter(event__id=request.GET['e'])
        elif 'q' in request.GET and request.GET['q'] != '':
            if request.user.event_user: context['voluntarys'] = Voluntary.objects.filter(name__icontains=request.GET['q'], event=request.user.event_user)
            else: context['voluntarys'] = Voluntary.objects.filter(name__icontains=request.GET['q'])
        elif user.type != 0:
            context['voluntarys'] = Voluntary.objects.filter(event=request.user.event_user)

        print(context)
        return render(request, 'voluntary/voluntary_manage.html', context)
    else:
        print(request.POST, "aa", request.FILES)
        get_param = request.GET.get('e') or request.POST.get('event') or ''
        redirect_url = f"{reverse('voluntary_manage')}?e={get_param}"
        if 'voluntary_delete' in request.POST:
            voluntary_id = request.POST.get('voluntary_delete')
            voluntary_delete = Voluntary.objects.get(id=voluntary_id)

            status = verificar_foto(str(voluntary_delete.photo))
            if status:
                voluntary_delete.photo.delete()
            voluntary_delete.delete()

            messages.success(request, f"{voluntary_delete.get_type_voluntary_display()} removido do sistema com sucesso!")

        elif 'voluntary_id' in request.POST:
            voluntary = get_object_or_404(Voluntary, id=request.POST.get("voluntary_id"))
            voluntary.name = request.POST.get('name')
            voluntary.registration = request.POST.get('registration')
            voluntary.type_voluntary = request.POST.get('type_voluntary')

            if request.user.is_staff:
                voluntary.admin = User.objects.get(id=request.POST.get('user'))
                voluntary.event = Event.objects.get(id=request.POST.get('event'))

            photo = request.FILES.get('photo')
            if photo:
                status = type_file(request, ['.png', '.jpg', '.jpeg'], photo,
                                   'A photo anexada não é do tipo png, jpg ou jpeg, considere converte-la em um desses tipos.')
                if status:
                    return redirect(redirect_url)

                if voluntary.photo:
                    status_photo = verificar_foto(str(voluntary.photo))
                    if status_photo:
                        voluntary.photo.delete()

                voluntary.photo = request.FILES.get('photo')

            voluntary.save()

        elif 'name' in request.POST:
            name = request.POST.get('name')
            registration = request.POST.get('registration')

            if not request.POST.get('type_voluntary'):
                messages.error(request, "Você não enviou todos os dados solicitados, confira e reenvie!")
                return redirect(redirect_url)

            type_voluntary = request.POST.get('type_voluntary')
            photo = request.FILES.get('photo')

            if photo:
                status = type_file(request, ['.png', '.jpg', '.jpeg'], photo,
                                   'A photo anexada não é do tipo png, jpg ou jpeg, considere converte-la em um desses tipos.')
                if status:
                    return redirect(redirect_url)

            if request.user.is_staff:
                event = Event.objects.get(id=request.POST.get('event'))
                admin = User.objects.get(id=request.POST.get('user'))
            else:
                admin = user
                event = request.user.event_user

            Voluntary.objects.create(
                type_voluntary=type_voluntary,
                name=name,
                registration=registration,
                admin=admin,
                photo=photo,
                event=event
            )
            messages.success(request, "Parabéns, você cadastrou mais um membro da comissão técnica!")
        return redirect(redirect_url)


@login_required(login_url="login")
@terms_accept_required
@permission_required('app.change_player_match', raise_exception=True)
@permission_required('app.add_player_match', raise_exception=True)
def players_in_teams(request, id):
        match = get_object_or_404(Match, id=id)
        team_match = Team_match.objects.filter(match=match)
        team_match_a = Team_match.objects.get(id=team_match[0].id)
        team_match_b = Team_match.objects.get(id=team_match[1].id)
        team_sport_a = Team_sport.objects.get(team=team_match_a.team, sport=team_match_a.match.sport, sexo=match.sexo)
        team_sport_b = Team_sport.objects.get(team=team_match_b.team, sport=team_match_b.match.sport, sexo=match.sexo)
        player_team_sport_a = Player_team_sport.objects.filter(team_sport=team_sport_a)
        player_team_sport_b = Player_team_sport.objects.filter(team_sport=team_sport_b)
        for i in player_team_sport_a:
            if not Player_match.objects.filter(player=i.player, match=match, team_match=team_match_a).exists():
                Player_match.objects.create(player=i.player, match=match, team_match=team_match_a)
        for i in player_team_sport_b:
            if not Player_match.objects.filter(player=i.player, match=match, team_match=team_match_b).exists():
                Player_match.objects.create(player=i.player, match=match, team_match=team_match_b)
        player_match_a = Player_match.objects.filter(match=match, team_match=team_match_a)
        player_match_b = Player_match.objects.filter(match=match, team_match=team_match_b)
        context = {
            'player_match_a':player_match_a,
            'player_match_b':player_match_b,
            'team_match_a':team_match_a,
            'team_match_b':team_match_b,
            
        }
        if request.method == "GET":
            return render(request, 'players_in_teams.html', context)
        else:
            if 'player_delete' in request.POST:
                if request.user.has_perm('app.delete_player_match'):
                    player_id = request.POST.get('player_delete')
                    player = Player_match.objects.get(id=player_id)
                    player.delete()
                else:
                    messages.error(request, "Você não tem permissão para remover o atleta da partida.")
            if 'team-a' in request.POST:
                for i in player_match_a:
                    number = request.POST.get(f'number_a_{i.id}')        
                    player = get_object_or_404(Player_match, id=i.id) 
                    if number != '': 
                        if int(number) >= 0:
                            player.player_number = number
                    player.save()
                messages.success(request, f"O número dos atletas do campus {team_match_a.team.get_campus_display()} foram adicionados/atualizados com sucesso!")
            if 'team-b' in request.POST:
                for i in player_match_b:
                    number = request.POST.get(f'number_b_{i.id}')          
                    player = get_object_or_404(Player_match, id=i.id)
                    if number != '': 
                        if int(number) >= 0:
                            player.player_number = number
                    player.save()
                messages.success(request, f"O número dos atletas do campus {team_match_b.team.get_campus_display()} foram adicionados/atualizados com sucesso!")
            return redirect('players_in_teams', match.id)

@login_required(login_url="login")
@terms_accept_required
@permission_required('app.change_player_match', raise_exception=True)
@permission_required('app.view_player_match', raise_exception=True)
def players_match(request, id):
    team_match = get_object_or_404(Team_match, id=id)
    player_match = Player_match.objects.filter(team_match=team_match)
    context = {
        'team_match': team_match,
        'player_match': player_match,
        
    }
    if request.method == "GET":
        return render(request, 'manage_players_match.html', context)
    else:
        try:
            players = request.POST.getlist('input-checkbox')
            select_action = request.POST.get('select-action')
            if 'select-action' in request.POST:
                if select_action == 'reserva':
                    for i in players:
                        player = get_object_or_404(Player, id=i)
                        player_match_status = Player_match.objects.get(player=player, team_match=team_match)
                        player_match_status.activity = 1
                        player_match_status.save()
                    return redirect('players_match', team_match.id)
                if select_action == 'titular':
                    for i in players:
                        player = get_object_or_404(Player, id=i)
                        player_match_status = Player_match.objects.get(player=player, team_match=team_match)
                        player_match_status.activity = 0
                        player_match_status.save()
                    return redirect('players_match', team_match.id)
                if select_action == 'excluir':
                    for i in players:
                        player = get_object_or_404(Player, id=i)
                        player_match = Player_match.objects.get(player=player, team_match=team_match)
                        player_match.delete()
                    return redirect('players_match', team_match.id)
            if 'player_match_delete' in request.POST:
                pass
                player_match_id = request.POST.get('player_match_delete')
                player_match = Player_match.objects.get(id=player_match_id)
                player_match.delete()
                return redirect('players_match', team_match.id)
        except (Player.DoesNotExist, Player_match.DoesNotExist):
            print('O jogador não foi encontrado :(')
        except Exception as e:
            print(f'Um erro inesperado aconteceu: {str(e)}')
        return redirect('players_match', team_match.id)

@login_required(login_url="login")
@terms_accept_required
@permission_required('app.change_player_match', raise_exception=True)
@permission_required('app.add_player_match', raise_exception=True)
@permission_required('app.view_player', raise_exception=True)
def add_players_match(request, id):
    team_match = get_object_or_404(Team_match, id=id)
    players = Player.objects.all()
    context = {
        'players': players,
        
    }
    if request.method == "GET":
        return render(request, 'add_players_match.html',context)
    else:
        try:
            player_id = request.POST.getlist('input-checkbox')
            print("IDs: ",player_id)
            for i in player_id:
                number = request.POST.get(f'number_{i}')
                if int(number) < 1:
                    messages.error(request, "Os números precisam ser maior que 1!")
                    return redirect('add_players_match', id)
                else:
                    print(request.POST.get(f'number_{i}'))              
                    print("ele ta continuando")
                    player = get_object_or_404(Player, id=i)
                    print(player)
                    player_match = Player_match.objects.create(match=team_match.match, team_match=team_match ,player=player, player_number=number)
                    print(player_match)
                    player_match.save()
                    print(player_match, "salvou")
        except ValueError:
            messages.error(request, "Você precisa informar os jogadores e seus respectivos números corretamente!")
            return redirect('add_players_match', id)

        return redirect('players_match', id)

@login_required(login_url="login")
@terms_accept_required
@permission_required('app.add_banner', raise_exception=True)
def banner_register(request):
    if request.method == "GET":
        return render(request, 'settings/banner_register.html')
    else:
        name = request.POST.get('name')
        image = request.FILES.get('banner')
        if not name or not image:
            messages.eror(request, "Você precisa preencher todas as informações!")
            return redirect('banner_register')
        Banner.objects.create(name=name,image=image)
        return redirect('banner_register')

@login_required(login_url="login")
@terms_accept_required
@permission_required('app.view_banner', raise_exception=True)
@permission_required('app.delete_banner', raise_exception=True)
@permission_required('app.change_banner', raise_exception=True)
def banner_manage(request):
    banner = Banner.objects.filter()
    if request.method == "GET":
        return render(request, 'settings/banner_manage.html',{'banner': banner})
    else:
        try:
            if 'banner_delete' in request.POST:
                banner_id = request.POST.get('banner_delete')
                banner_delete = Banner.objects.get(id=banner_id)
                banner_delete.delete()
                return redirect('banner_manage')
            if 'banner_update' in request.POST:
                banner_id = request.POST.get('banner_update')
                banner = Banner.objects.get(id=banner_id)
                if banner.status == 0: banner.status = 1
                elif banner.status == 1: 
                    banner.status = 0
                    if Banner.objects.filter(status=1):
                        banner2 = Banner.objects.filter(status=0)
                        for i in banner2:
                            i.status = 1
                            i.save()
                banner.save()
                return redirect('banner_manage')
        except Exception as e: messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
        return redirect('banner_manage')

@login_required(login_url="login")
def upload_document(request):
    termo = User.objects.get(id=request.user.id)
    if not request.user.event_user.general_need_authorization:
        return redirect('boss_data')
    if termo.document:
        return redirect('boss_data')
    
    if request.method == 'POST':
        document = request.FILES.get('document')
        if document:
            termo.document = document
            if document: 
                status = type_file(request, ['.pdf','.png','.jpg','.jpeg'], document, 'O documento anexado precisa ser do tipo pdf, png, jpg ou jpeg.')
                if status: return redirect('upload_document')
            termo.save()
            return redirect('boss_data')

    return render(request, 'terms/terms_use_upload.html')


@login_required(login_url="login")
def boss_data(request):
    termo = User.objects.get(id=request.user.id)
    if request.user.event_user.general_need_authorization:
        if not termo.document:
            return redirect('upload_document')
    if termo.email and termo.photo and termo.telefone:
        return redirect('terms_use')

    if request.method == 'POST':
        password = request.POST.get('password')
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        photo = request.FILES.get('photo')

        if password and photo and email and phone and name:
            termo.first_name = name
            termo.email = email
            termo.telefone = phone
            if photo: 
                status = type_file(request, ['.png','.jpg','.jpeg'], photo, 'A photo anexada não é do tipo png, jpg ou jpeg, considere converte-la em um desses tipos.')
                if status: return redirect('boss_data')
            termo.photo = photo
            termo.set_password(password)
            termo.save()
            update_session_auth_hash(request, termo)
            return redirect('terms_use')
        else:
            messages.error(request, 'Você precisa preencher todas as informações!')
            return redirect('boss_data')

    return render(request, 'terms/terms_use_data.html')


@login_required(login_url="login")
def terms_use(request):
    user = User.objects.get(id=request.user.id)
    if request.user.event_user.general_need_authorization:
        if not user.document:
            return redirect('upload_document')
    if not (user.email and user.photo and user.telefone and user.first_name):
        return redirect('boss_data')
    if user.accepted:
        return redirect('Home')

    if request.method == 'POST':
        if request.POST.get('accept') == 'on':
            user.accepted = True
            user.accepted_at = timezone.now()
            user.save()

            Voluntary.objects.create(
                name=user.first_name,
                photo=user.photo,
                unit=user.unit,
                admin=request.user,
                event=user.event_user,
                type_voluntary=2
            )

            return redirect('Home')

    return render(request, 'terms/terms_use.html' , {'termo': user})

@login_required(login_url="login")  
def settings(request):
    return render(request, 'settings.html')

@login_required(login_url="login")  
def settings_new(request):
    if request.POST:
        if 'banner_delete' in request.POST:
            banner = Banner.objects.get(id=request.POST.get('banner_delete'))
            banner.image.delete()
            banner.delete()
        elif 'attachments_delete' in request.POST:
            attachments = Attachments.objects.get(id=request.POST.get('attachments_delete'))
            attachments.file.delete()
            attachments.delete()
        elif 'statement_delete' in request.POST:
            statement = Statement.objects.get(id=request.POST.get('statement_delete'))
            statement.image.delete()
            statement.delete()
    context = {
        'banners': Banner.objects.all(),
        'attachments': Attachments.objects.all(),
        'statement': Statement.objects.all(),
    }
    return render(request, 'settings_new.html', context)

@login_required(login_url="login")
@terms_accept_required
@permission_required('app.add_statement', raise_exception=True)
def statement_register(request):
    if request.method == "GET":
        return render(request, 'settings/statement_register.html')
    else:
        name = request.POST.get('name')
        image = request.FILES.get('image')
        if not name or not image:
            messages.eror(request, "Você precisa preencher todas as informações!")
            return redirect('statement_register')
        Statement.objects.create(name=name,image=image)
        return redirect('statement_manage')

@login_required(login_url="login")
@terms_accept_required
@permission_required('app.view_statement', raise_exception=True)
@permission_required('app.delete_statement', raise_exception=True)
@permission_required('app.view_statement_user', raise_exception=True)
@permission_required('app.delete_statement_user', raise_exception=True)
def statement_manage(request):
    statement = Statement.objects.filter()
    statement_user = Statement_user.objects.all().order_by('statement')
    if request.method == "GET":
        return render(request, 'settings/statement_manage.html',{'statement': statement,'statement_user': statement_user})
    else:
        try:
            if 'statement_delete' in request.POST:
                statement_id = request.POST.get('statement_delete')
                statement_delete = Statement.objects.get(id=statement_id)
                statement_delete.delete()
                return redirect('statement_manage')
            elif 'statement_user_delete' in request.POST:
                statement_user_id = request.POST.get('statement_user_delete')
                statement_user_delete = Statement_user.objects.get(id=statement_user_id)
                statement_user_delete.delete()
                return redirect('statement_manage')
        except Exception as e: messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
        return redirect('statement_manage')

@login_required(login_url="login")
@permission_required('app.view_terms_use', raise_exception=True)
def chefe_manage(request):
    if request.method == "GET":
        terms = Terms_Use.objects.all()
        return render(request, 'settings/chefe_manage.html',{'terms': terms})
    else:
        try:
            if request.user.has_perm('app.delete_terms_use'):
                terms_delete = request.POST.get('terms_delete')
                term_del = Terms_Use.objects.get(id=terms_delete)
                term_del.delete()
                messages.success(request, "Excluido com sucesso!")
            else:
                messages.error(request, "Você não tem permissão para remover.")
            return redirect('chefe_manage')
        except Exception as e: messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
        return redirect('chefe_manage')

@login_required(login_url="login")
@permission_required('app.view_help', raise_exception=True)
def faq_manage(request):
    if request.method == "GET":
        help = Help.objects.all()
        return render(request, 'settings/faq_manage.html',{'help': help})
    else:
        try:
            if request.user.has_perm('app.delete_help'):
                faq_delete = request.POST.get('faq_delete')
                faq = Help.objects.get(id=faq_delete)
                faq.delete()
                messages.success(request, "Excluido com sucesso!")
            else:
                messages.error(request, "Você não tem permissão para remover.")
            return redirect('faq_manage')
        except Exception as e: messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
        return redirect('faq_manage')
    
@login_required(login_url="login")
@permission_required('app.view_attachments', raise_exception=True)
def anexo_manage(request):
    if request.method == "GET":
        atack = Attachments.objects.all().order_by('-id')
        return render(request, 'settings/anexo_manage.html',{'atack': atack})
    else: 
        try:
            if request.user.has_perm('app.delete_attachments'):
                atack_delete = request.POST.get('atack_delete')
                atack = Attachments.objects.get(id=atack_delete)
                atack.delete()
                messages.success(request, "Excluido com sucesso!")
            else:
                messages.error(request, "Você não tem permissão para remover.")
            return redirect('anexo_manage')
        except Exception as e: messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
        return redirect('anexo_manage')
    
@login_required(login_url="login")
@permission_required('app.view_settings_access', raise_exception=True)
def enrollment_manage(request):
    if request.method == "GET":
        date_list = Settings_access.objects.all().order_by('-id')
        return render(request, 'settings/enrollment_manage.html',{'date_list': date_list})
    else:
        try:
            if request.user.has_perm('app.delete_settings_access'):
                date_delete = request.POST.get('date_delete')
                settings_access = Settings_access.objects.get(id=date_delete)
                settings_access.delete()
                messages.success(request, "Excluido com sucesso!")
            else:
                messages.error(request, "Você não tem permissão para remover.")
            return redirect('enrollment_manage')
        except Exception as e: messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
        return redirect('enrollment_manage')
    
@login_required(login_url="login")
@permission_required('app.add_help', raise_exception=True)
def faq_register(request):
    if request.method == "GET":
        return render(request, 'settings/faq_register.html')
    else:
        try:
            title_faq = request.POST.get('title_faq')
            details_faq = request.POST.get('details_faq')
            Help.objects.create(title=title_faq, description=details_faq)
            messages.success(request, "Parabéns, foi cadastrado com sucesso!")
            return redirect('faq_manage')
        except Exception as e: messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
        return redirect('faq_manage')
    
@login_required(login_url="login") 
@permission_required('app.add_attachments', raise_exception=True)
def anexo_register(request):
    if request.method == "GET":
        return render(request, 'settings/anexo_register.html')
    else:
        try:
            title_atack = request.POST.get('title_atack')
            file_atack = request.FILES.get('file_atack')
            print(file_atack)
            print(request.POST, request.FILES)
            Attachments.objects.create(name=title_atack, user=request.user, file=file_atack)
            messages.success(request, "Parabéns, foi cadastrado com sucesso!")
            return redirect('anexo_manage')
        except Exception as e: messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
        return redirect('anexo_manage')
    
@login_required(login_url="login")
@permission_required('app.add_settings_access', raise_exception=True)
def enrollment_register(request):
    if request.method == "GET":
        return render(request, 'settings/enrollment_register.html')
    else:
        try:
            start = request.POST.get('date_start')
            end = request.POST.get('date_end')
            print(f'i: {start} - f: {end}')
            Settings_access.objects.create(start=start, end=end)
            messages.success(request, "Parabéns, foi cadastrado com sucesso!")
            return redirect('enrollment_manage')
        except Exception as e: messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
        return redirect('enrollment_manage')

@login_required(login_url="login")
@permission_required('app.view_match', raise_exception=True)
@permission_required('app.add_point', raise_exception=True)
@permission_required('app.add_penalties', raise_exception=True)
def scoreboard(request, event_id):  
    match_event = Event.objects.get(id=event_id)
    referee = Voluntary.objects.filter(event=match_event, type_voluntary=6)
    types_referee = Type_referee.choices
    time_now = time.strftime("%H:%M:%S", time.localtime())
    
    if Match.objects.filter(status=1, event=match_event):
        time_now2 = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        match = Match.objects.get(status=1, event=match_event)
        match_referee = Match_referee.objects.filter(match=match)
        matches = Match.objects.filter(status=0, event=match_event)
        team_match_all = Team_match.objects.filter(team__event=match_event)
        team_matchs = Team_match.objects.filter(match=match)
        team_match_a = team_matchs[0]
        team_match_b = team_matchs[1]
        banners = Banner.objects.all()
        players_match_a = Player_match.objects.filter(team_match=team_match_a)
        players_match_b = Player_match.objects.filter(team_match=team_match_b)  
        team_sport_a = Team_sport.objects.get(team=team_match_a.team, sport__sport=team_match_a.match.sport, sexo=match.sexo)
        team_sport_b = Team_sport.objects.get(team=team_match_b.team, sport__sport=team_match_b.match.sport, sexo=match.sexo)
        player_team_sport_a = Player_team_sport.objects.filter(team_sport=team_sport_a)
        player_team_sport_b = Player_team_sport.objects.filter(team_sport=team_sport_b)
        seconds, status = generate_timer(match)
        for i in player_team_sport_a:
            if not Player_match.objects.filter(player=i.player, match=match, team_match=team_match_a).exists():
                Player_match.objects.create(player=i.player, match=match, team_match=team_match_a)
        for i in player_team_sport_b:
            if not Player_match.objects.filter(player=i.player, match=match, team_match=team_match_b).exists():
                Player_match.objects.create(player=i.player, match=match, team_match=team_match_b)
        for i in players_match_a:
            if not Player_team_sport.objects.filter(player=i.player, team_sport=team_sport_a).exists():
                i.delete()
        for i in players_match_b:
            if not Player_team_sport.objects.filter(player=i.player, team_sport=team_sport_b).exists():
                i.delete()
        if match.sport == 0:
            point_a = Point.objects.filter(team_match=team_match_a).count() - Point.objects.filter(point_types=2,team_match=team_match_a).count()
            point_b = Point.objects.filter(team_match=team_match_b).count() - Point.objects.filter(point_types=2,team_match=team_match_b).count()
        else:      
            point_a = Point.objects.filter(team_match=team_match_a).count()
            point_b = Point.objects.filter(team_match=team_match_b).count()
    if request.method == "GET":
        context = {
            'event': match_event,
            'point_types': Point_types.choices,
            'penalities_types': Type_penalties.choices,
            'match': match,
            'team_match_a': team_match_a,
            'team_match_b': team_match_b,
            'point_a': point_a,
            'point_b': point_b,
            'players_match_a': players_match_a,
            'players_match_b': players_match_b,
            'points': Point.objects.all(),
            'activity_types': Activity.choices,
            'occurrence': Occurrence.objects.order_by('-id')[:7],
            'seconds': seconds,
            'status': status,
            'banners': banners,
            'matches': matches,
            'team_match_all': team_match_all,
            'detailed': Detailed.choices,
            'referee': referee,
            'types_referee': types_referee,
            'match_referee': match_referee,

        }
        print("context")
        return render(request, 'scoreboard.html', context)
    else:
        print(request.POST)
        if 'detailed' in request.POST:
            print("det")
            detailed = int(request.POST.get('detailed'))
            match.detailed = detailed
            match.save()
        elif 'team-a' in request.POST:
            for i in players_match_a:
                number = request.POST.get(f'number_a_{i.id}') 
                activity = request.POST.get(f'activity_a_{i.id}')        
                player = get_object_or_404(Player_match, id=i.id) 
                if number != '': 
                    if int(number) >= 0:
                        player.player_number = number
                        player.activity = int(activity)
                player.save()
            messages.success(request, f"Dados atualizados!")
        elif 'team-b' in request.POST:
            for i in players_match_b:
                number = request.POST.get(f'number_b_{i.id}')   
                activity = request.POST.get(f'activity_b_{i.id}')           
                player = get_object_or_404(Player_match, id=i.id)
                if number != '': 
                    if int(number) >= 0:
                        player.player_number = number
                        player.activity = int(activity)
                player.save()
            messages.success(request, f"Dados atualizados!")
        elif 'assistance' in request.POST:
            point = Point.objects.get(id=request.POST.get('point'))
            player = Player_match.objects.get(id=request.POST.get('player_id'))
            Assistance.objects.create(assis_to=point, player=player)
        elif 'banner' in request.POST:
            banner = Banner.objects.get(id=request.POST.get('banner'))
            if banner.status == 0: 
                banner.status = 1
            else: 
                if Banner.objects.filter(status=0):
                    banners = Banner.objects.filter(status=0)
                    for i in banners:
                        i.status = 1
                        i.save()
                banner.status = 0
            banner.save()

        elif 'replacement_init' in request.POST:
            player_init = get_object_or_404(Player_match, id=request.POST.get("replacement_init"))
            player_end = get_object_or_404(Player_match, id=request.POST.get("replacement_end"))
            player_init.activity = int(player_end.activity)
            player_end.activity = 1
            Replacement.objects.create(team_match=player_init.team_match, player_entry=player_init, player_exit=player_end)
            player_init.save(), player_end.save()
        elif 'referee' in request.POST:
            referee = Voluntary.objects.get(id=request.POST.get("referee"))
            referee_type = int(request.POST.get("type_referee"))
            Match_referee.objects.create(match=match, referee=referee, role=referee_type)
        elif 'color_a' in request.POST or 'color_b' in request.POST:
            print(request)
            if request.POST.get("color_a"):
                team_match_a.team.color = str(request.POST.get("color_a"))
                team_match_a.team.save()
            if request.POST.get("color_b"):
                team_match_b.team.color = str(request.POST.get("color_b"))
                team_match_b.team.save()
        elif 'observations' in request.POST:
            match.observations = request.POST.get("observations")
            match.save()
        elif 'penalties' in request.POST:
            penalties_type = request.POST.get('penalties')
            player_match = Player_match.objects.get(id=request.POST.get('player_penalties'))
            penalties = Penalties.objects.create(player=player_match.player, type_penalties=int(penalties_type), team_match=player_match.team_match)
            penalties.save()
            
            details = f"{player_match.player.name} recebeu {penalties.get_type_penalties_display().lower()}"
            Occurrence.objects.create(name=penalties.get_type_penalties_display(), details=details, match=match)
        elif 'team-a-point' in request.POST:
            if match.sport == 0: type = 0
            else: type = 1
            if request.POST.get("team-a-point") == "+1":
                if request.POST.get("player-a-point"):
                    player = Player.objects.get(id=request.POST.get("player-a-point"))
                    point = Point.objects.create(team_match=team_match_a, player=player, point_types=type)
                    details = f"{player.name} fez um {point.get_point_types_display().lower()}"
                    Occurrence.objects.create(name=point.get_point_types_display(), details=details, match=match)
                else: 
                    point = Point.objects.create(team_match=team_match_a, point_types=type)
                point.save()
            elif Point.objects.filter(team_match=team_match_a, point_types=type).exists(): 
                point = Point.objects.filter(team_match=team_match_a, point_types=type).last().delete()
        elif 'team-b-point' in request.POST:
            if match.sport == 0: type = 0
            else: type = 1
            if request.POST.get("team-b-point") == "+1":
                if request.POST.get("player-b-point"):
                    player = Player.objects.get(id=request.POST.get("player-b-point"))
                    point = Point.objects.create(team_match=team_match_b, player=player, point_types=type)
                    details = f"{player.name} fez um {point.get_point_types_display().lower()}"
                    Occurrence.objects.create(name=point.get_point_types_display(), details=details, match=match)
                else: 
                    point = Point.objects.create(team_match=team_match_b, point_types=type)
                point.save()
            elif Point.objects.filter(team_match=team_match_b, point_types=type).exists(): 
                point = Point.objects.filter(team_match=team_match_b, point_types=type).last().delete()
        elif 'team-a-aces' in request.POST:
            if request.POST.get("team-a-aces") == "+1":
                if request.POST.get("player-a-point"):
                    player = Player.objects.get(id=request.POST.get("player-a-point"))
                    point = Point.objects.create(team_match=team_match_a, player=player, point_types=2)
                    details = f"{player.name} fez um {point.get_point_types_display().lower()}"
                    Occurrence.objects.create(name=point.get_point_types_display(), details=details, match=match)
                else: 
                    point = Point.objects.create(team_match=team_match_a, point_types=2)
                point.save()
            elif Point.objects.filter(team_match=team_match_a, point_types=2).exists(): 
                point = Point.objects.filter(team_match=team_match_a, point_types=2).last().delete()
        elif 'team-b-aces' in request.POST:
            if request.POST.get("team-b-aces") == "+1":
                if request.POST.get("player-b-point"):
                    player = Player.objects.get(id=request.POST.get("player-b-point"))
                    point = Point.objects.create(team_match=team_match_b, player=player, point_types=2)
                    details = f"{player.name} fez um {point.get_point_types_display().lower()}"
                    Occurrence.objects.create(name=point.get_point_types_display(), details=details, match=match)
                else: 
                    point = Point.objects.create(team_match=team_match_b, point_types=2)
                point.save()
            elif Point.objects.filter(team_match=team_match_b, point_types=2).exists(): 
                point = Point.objects.filter(team_match=team_match_b, point_types=2).last().delete()
        elif 'volley_new' in request.POST:
            print("Bora pro vô léi")
            if match.volley_match and match.status == 1:
                volley_match = Volley_match.objects.get(status=1)
                match.status = 2
                if point_a > point_b:
                    match.Winner_team = team_match_a.team
                    volley_match.sets_team_a += 1
                elif point_b > point_a:
                    match.Winner_team = team_match_b.team
                    volley_match.sets_team_b += 1
                match.save()
                volley_match.save()
                print("CRIANDO NOVO SET")
                print("1:: ",volley_match)
                new_match = Match.objects.create(sport=1, sexo=match.sexo, event=match.event, status=5, volley_match=volley_match, time_match=time_now2)
                if match.location:
                    new_match.location = match.location
                if match.group_phase:
                    new_match.group_phase = match.group_phase
                print(new_match)
                new_match.save()
                team_a_match = Team_match.objects.create(match=new_match, team=team_match_a.team)
                team_a_match.save()
                print("2:: ",team_a_match)
                team_b_match = Team_match.objects.create(match=new_match, team=team_match_b.team)
                team_b_match.save()
                print("3:: ",team_b_match)
                new_match.status = 1
                new_match.save()
                for i in players_match_a:
                    player_match = Player_match.objects.create(match=new_match, team_match=team_a_match ,player=i.player, player_number=i.player_number, activity=i.activity)
                    player_match.save()
                for i in players_match_b:
                    player_match = Player_match.objects.create(match=new_match, team_match=team_b_match ,player=i.player, player_number=i.player_number, activity=i.activity)
                    player_match.save()
                print("criado com sucesso")
                return redirect('scoreboard', match_event.id)
            else:
                messages.error(request, 'OS SETS SÓ PODEM SER CRIADOS EM ESPORTES QUE NECESSITAM DELE. EX: VOLEIBOL')
                print("O sets só podem ser criados em esportes que necessitam dele. Ex: Voleibol")
                return redirect('scoreboard', match_event.id)
            
        elif 'finally' in request.POST:
            if match.time_start and not match.time_end:
                messages.error(request, "Antes de finalizar a partida e iniciar outra você precisa primeiro parar o cronometro!")
                return redirect('scoreboard', match_event.id)
            print("chegou na primeira parte")
            if Volley_match.objects.filter(status=1) or match.sport in [1,2]:
                volley_match = get_object_or_404(Volley_match, status=1)
                match.status = 2
                match.detailed = 3
                if point_a > point_b:
                    match.Winner_team = team_match_a.team
                    volley_match.sets_team_a += 1
                elif point_b > point_a:
                    match.Winner_team = team_match_b.team
                    volley_match.sets_team_b += 1
                match.save()
                volley_match.status= 2 
                volley_match.save()
            else:
                match.status = 2
                if point_a > point_b:
                    match.Winner_team = team_match_a.team
                elif point_b > point_a:
                    match.Winner_team = team_match_b.team
                match.save()

            return redirect('games')
        elif 'match_new' in request.POST:
            if match.time_start and not match.time_end:
                messages.error(request, "Antes de finalizar a partida e iniciar outra você precisa primeiro parar o cronometro!")
                return redirect('scoreboard', match_event.id)
            print("chegou na primeira parte")
            next_match_id = request.POST.get('match_new')
            next_match = Match.objects.get(id=next_match_id)
            if match.sport in [1,2]:
                print("A partida anterios é de vollei")
                volley_match = get_object_or_404(Volley_match, status=1)
                print("volley: ",volley_match)
                print("mudando estatus da partida")
                match.status = 2
                match.detailed = 3
                if point_a > point_b:
                    match.Winner_team = team_match_a.team
                    volley_match.sets_team_a += 1
                elif point_b > point_a:
                    match.Winner_team = team_match_b.team
                    volley_match.sets_team_b += 1
                match.save()
                print("mUdeii:")
                print(volley_match.status)
                volley_match.status = 2 
                volley_match.save()
                print("Foi finalizada de vez!")
            else:
                print("A partida anterios é qualquer uma")
                match.status = 2
                match.detailed = 3
                if point_a > point_b:
                    match.Winner_team = team_match_a.team
                elif point_b > point_a:
                    match.Winner_team = team_match_b.team
                match.save()
                print("Foi finalizada!", match.get_status_display())
            if next_match.volley_match:
                print("A proxima é de vollei tb")
                volley_match = Volley_match.objects.get(id=next_match.volley_match.id)
                print(volley_match)
                print(volley_match.status)
                volley_match.status = 1
                volley_match.save()
            else:
                print("A proxima é aleatoria")
            print(next_match, next_match.get_status_display())
            team_matchs = Team_match.objects.filter(match=next_match)
            if team_matchs[0] and team_matchs[1]:
                if team_matchs[0].team.photo and team_matchs[1].team.photo:
                    next_match.status = 1
                    next_match.save()
                    return redirect('scoreboard', match_event.id)
            else:
                messages.error(request, "É necessário que tenha 2 times!")
                next_match.status = 3
                next_match.save()
                print(next_match)
                return redirect('scoreboard', match_event.id)
        elif 'time_init' in request.POST:
            if match.time_start and match.time_end:
                print("O cronometro já finalizou!")
                return redirect('scoreboard', match_event.id)
            
            elif match.time_start:
                if Time_pause.objects.filter(match=match):
                    pause = Time_pause.objects.filter(match=match).last()
                    if pause.start_pause and not pause.end_pause:
                        pause.end_pause = time_now
                        pause.save()
                        match.detailed = 1
                        match.save()
                        return redirect('scoreboard', match_event.id)                
                    else:
                        pause_time = Time_pause.objects.create(start_pause=time_now,match=match)
                        pause_time.save()
                        match.detailed = 2
                        match.save()
                        return redirect('scoreboard', match_event.id)
                else:
                    pause_time = Time_pause.objects.create(start_pause=time_now,match=match)
                    pause_time.save()
                    match.detailed = 2
                    match.save()
                    print(pause_time)
                    return redirect('scoreboard', match_event.id)
            
            else:
                match.time_start = time_now
                match.save()
                match.detailed = 1
                match.save()
                return redirect('scoreboard', match_event.id)
                
        elif 'time_stop' in request.POST:
            if match.time_start and match.time_end:
                print("O cronometro foi finalizado!")
                return redirect('scoreboard', match_event.id)
            elif match.time_start:
                if Time_pause.objects.filter(match=match).last():
                    pause = Time_pause.objects.filter(match=match).last()
                    if pause.start_pause and not pause.end_pause:
                        pause.end_pause = time_now
                        pause.save()
                match.time_end = time_now
                match.detailed = 3
                match.save()
                return redirect('scoreboard', match_event.id)
            else:
                print("o cronometro precisa ser iniciado, para ser finalizado!")
                return redirect('scoreboard', match_event.id)
        return redirect('scoreboard', match_event.id)
    
def scoreboard_public(request, event_id):
    try:
        event = Event.objects.get(id=event_id)
        match = None
        if Volley_match.objects.filter(status=1, event=event).exists():
            volley_match = Volley_match.objects.get(status=1, event=event)
            if Match.objects.filter(volley_match=volley_match, status=1, event=event).exists():
                match = Match.objects.get(volley_match=volley_match, status=1, event=event)
        elif Match.objects.filter(status=1, event=event):
            match = Match.objects.get(status=1, event=event)
        if match:
            print("match sim")
            seconds, status = generate_timer(match)
            team_matchs = Team_match.objects.filter(match=match)

            if match.volley_match:
                if (match.volley_match.sets_team_a + match.volley_match.sets_team_b) % 2 == 0:
                    team_match_a = team_matchs[0]
                    team_match_b = team_matchs[1]
                    sets_a = match.volley_match.sets_team_a
                    sets_b = match.volley_match.sets_team_b
                else:
                    team_match_a = team_matchs[1]
                    team_match_b = team_matchs[0]
                    sets_b = match.volley_match.sets_team_a
                    sets_a = match.volley_match.sets_team_b
                ball_sport = static('images/ball-of-volley.png')
            else:
                team_match_a = team_matchs[0]
                team_match_b = team_matchs[1]
                if match.sport == 3: ball_sport = static('images/ball-of-handball.png')
                else: ball_sport = static('images/ball-of-futsal.png')

            players_match_a = Player_match.objects.filter(team_match=team_match_a)
            players_match_b = Player_match.objects.filter(team_match=team_match_b)
            if match.sport == 0:
                point_a = Point.objects.filter(team_match=team_match_a).count() - Point.objects.filter(point_types=2,team_match=team_match_a).count()
                point_b = Point.objects.filter(team_match=team_match_b).count() - Point.objects.filter(point_types=2,team_match=team_match_b).count()
            else:
                point_a = Point.objects.filter(team_match=team_match_a).count()
                point_b = Point.objects.filter(team_match=team_match_b).count()
            card_a = Penalties.objects.filter(type_penalties=0, team_match=team_match_a).count() + Penalties.objects.filter(type_penalties=1, team_match=team_match_a).count()
            card_b = Penalties.objects.filter(type_penalties=0, team_match=team_match_b).count() + Penalties.objects.filter(type_penalties=1, team_match=team_match_b).count()
            lack_a = Penalties.objects.filter(type_penalties=2,team_match=team_match_a).count()
            lack_b = Penalties.objects.filter(type_penalties=2,team_match=team_match_b).count()
            
            occurrence = Occurrence.objects.filter(match=match).order_by('-datetime')[:10]
            context = {
                'match': match,
                'team_match_a':team_match_a,
                'team_match_b':team_match_b,
                'players_match_a':players_match_a,
                'players_match_b':players_match_b,
                'point_a':point_a,
                'point_b':point_b,
                'lack_a':lack_a,
                'lack_b':lack_b,
                'ball_sport': ball_sport,
                'card_a': card_a,
                'card_b': card_b,
                'events': occurrence,
                'event':event,
            }
            print(context)
            if match.volley_match:
                context['aces_a'] = Point.objects.filter(point_types=2,team_match=team_match_a).count()
                context['aces_b'] = Point.objects.filter(point_types=2,team_match=team_match_b).count()
                context['sets_a'] = sets_a
                context['sets_b'] = sets_b
            else:
                context['seconds'] = seconds
                context['status'] = status
            if match.sport == 0:
                context['penalties_a'] = Point.objects.filter(point_types=2,team_match=team_match_a).count()
                context['penalties_b'] = Point.objects.filter(point_types=2,team_match=team_match_b).count()
            return render(request, 'public/scoreboard_public.html', context)
        else:
            print("match não")
            return render(request, 'public/scoreboard_public.html',{'event':event})
    except Exception as e:
        messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
        return render(request, 'public/scoreboard_public.html',{'event':event})

import qrcode
from io import BytesIO
import base64   

def scoreboard_projector(request, event_id):
    try:
        event = Event.objects.get(id=event_id)
        url = request.get_host()
        print(url)

        qr = qrcode.make(url)

        buffer = BytesIO()
        qr.save(buffer, format='PNG')
        buffer.seek(0)

        img_base64 = base64.b64encode(buffer.getvalue()).decode()

        if Volley_match.objects.filter(status=1, event=event):
            print("PARTIDA DE VOLEI")
            volley_match = Volley_match.objects.get(status=1, event=event)
            match = Match.objects.filter(volley_match=volley_match.id, event=event).last()
            team_matchs = Team_match.objects.filter(match=match)
            team_match_a = team_matchs[0]
            team_match_b = team_matchs[1]
            if (match.volley_match.sets_team_a + match.volley_match.sets_team_b) % 2 == 0:
                print("par")
                teammatch1 = team_match_a
                teammatch2 = team_match_b
                sets_1 = match.volley_match.sets_team_a
                sets_2 = match.volley_match.sets_team_b
            else:
                print("impar")
                teammatch1 = team_match_b
                teammatch2 = team_match_a
                sets_1 = match.volley_match.sets_team_b
                sets_2 = match.volley_match.sets_team_a
            if Banner.objects.filter(status=0): 
                banner_score = Banner.objects.get(status=0).image.url
                banner_bol = True
            else: 
                banner_score = static('images/logo-morea.svg')
                banner_bol = False
            players_match_a = Player_match.objects.filter(team_match=teammatch1)
            players_match_b = Player_match.objects.filter(team_match=teammatch2)
            point_a = Point.objects.filter(team_match=teammatch1).count()
            point_b = Point.objects.filter(team_match=teammatch2).count()
            print(point_a, point_b)
            aces_a = Point.objects.filter(point_types=2,team_match=teammatch1).count()
            aces_b = Point.objects.filter(point_types=2,team_match=teammatch2).count()
            lack_a = Penalties.objects.filter(type_penalties=2,team_match=teammatch1).count()
            lack_b = Penalties.objects.filter(type_penalties=2,team_match=teammatch2).count()
            card_a = Penalties.objects.filter(type_penalties=0,team_match=teammatch1).count() + Penalties.objects.filter(type_penalties=1,team_match=teammatch1).count()
            card_b = Penalties.objects.filter(type_penalties=0,team_match=teammatch2).count() + Penalties.objects.filter(type_penalties=1,team_match=teammatch2).count()
            occurrence = Occurrence.objects.filter()
            name_scoreboard = 'Sets'
            ball_sport = static('images/ball-of-volley.png')
            if match.sexo == 1: 
                img_sexo = static('images/icon-female.svg')
                sexo_color = '#ff32aa' 
            else: 
                img_sexo = static('images/icon-male.svg')
                sexo_color = '#3a7bd5'
            context = {
                'match': match,
                'time_sets_a': sets_1,
                'sets_b': sets_2,
                'team_match_a':teammatch1,
                'team_match_b':teammatch2,
                'players_match_a':players_match_a,
                'players_match_b':players_match_b,
                'point_a':point_a,
                'point_b':point_b,
                'lack_a':lack_a,
                'lack_b':lack_b,
                'img_sexo':img_sexo,
                'sexo_color': sexo_color,
                'ball_sport': ball_sport,
                'aces_a': aces_a,
                'aces_b': aces_b,
                'card_a':card_a,
                'card_b':card_b,
                'colorA': teammatch1.team.color,
                'colorB': teammatch2.team.color,
                'events': occurrence,
                'banner_score':banner_score,
                'banner_bol':banner_bol,
                'sexo_text':match.get_sexo_display(),
                'name_scoreboard': name_scoreboard,
                'qrcode': img_base64,
                'url': url,
                'event': event,
                
                
            }
            print(context)
            return render(request, 'public/scoreboard_projector.html', context)
            
        elif Match.objects.filter(status=1, event=event):
            print("sport aleatorio")
            match = Match.objects.get(status=1, event=event)
            occurrence = Occurrence.objects.filter(match=match)
            team_matchs = Team_match.objects.filter(match=match)
            team_match_a = team_matchs[0]
            team_match_b = team_matchs[1]
            players_match_a = Player_match.objects.filter(team_match=team_match_a)
            players_match_b = Player_match.objects.filter(team_match=team_match_b)
            point_a = Point.objects.filter(team_match=team_match_a).count() - Point.objects.filter(point_types=2,team_match=team_match_a).count()
            point_b = Point.objects.filter(team_match=team_match_b).count() - Point.objects.filter(point_types=2,team_match=team_match_b).count()
            penalties_a = Point.objects.filter(point_types=2,team_match=team_match_a).count()
            penalties_b = Point.objects.filter(point_types=2,team_match=team_match_b).count()
            lack_a = Penalties.objects.filter(type_penalties=2,team_match=team_match_a).count()
            lack_b = Penalties.objects.filter(type_penalties=2,team_match=team_match_b).count()
            card_a = Penalties.objects.filter(type_penalties=0,team_match=team_match_a).count() + Penalties.objects.filter(type_penalties=1,team_match=team_match_a).count()
            card_b = Penalties.objects.filter(type_penalties=0,team_match=team_match_b).count() + Penalties.objects.filter(type_penalties=1,team_match=team_match_b).count()
            seconds, status = generate_timer(match)
            name_scoreboard = 'Tempo'
            if match.sport == 3:
                ball_sport = static('images/ball-of-handball.png')
            else:
                ball_sport = static('images/ball-of-futsal.png')
            if Banner.objects.filter(status=0): 
                banner_score = Banner.objects.get(status=0).image.url
                banner_bol = True
            else: 
                banner_score = static('images/logo-morea.svg')
                banner_bol = False
            context = {
                'match': match,
                'events':occurrence,
                'time_sets_a': "00:00",
                'status': status,
                'seconds': seconds,
                'team_match_a':team_match_a,
                'team_match_b':team_match_b,
                'players_match_a':players_match_a,
                'players_match_b':players_match_b,
                'point_a':point_a,
                'point_b':point_b,
                'penalties_a':penalties_a,
                'penalties_b':penalties_b,
                'lack_a':lack_a,
                'lack_b':lack_b,
                'ball_sport': ball_sport,
                'aces_a': 0,
                'aces_b': 0,
                'colorA': team_match_a.team.color,
                'colorB': team_match_b.team.color,
                'banner_score':banner_score,
                'banner_bol':banner_bol,
                'card_a': card_a,
                'card_b': card_b,
                'sexo_text':match.get_sexo_display(),
                'name_scoreboard': name_scoreboard,
                'qrcode': img_base64,
                'url': url,
                'event': event,
            }
            return render(request, 'public/scoreboard_projector.html', context)
        else:
            return render(request, 'public/scoreboard_projector.html', {'qrcode': img_base64, 'event': event, 'url': url, 'colorA': "#FF0000", 'colorB': "#0000FF"})
    except Exception as e:
        messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
        return render(request, 'public/scoreboard_projector.html')

@login_required(login_url="login")
@terms_accept_required
def generator_badge(request):
    current_get_params = request.GET.urlencode()
    user = User.objects.get(id=request.user.id)
    if request.method == "GET":
        context = {}
        if request.user.is_staff or request.user.type == 0:
            context['events'] = Event.objects.all()
            if 'e' in request.GET and request.GET.get('e') != '':
                event = Event.objects.get(id=request.GET.get('e'))
                context['event'] = event 
                context['teams'] = Team.objects.filter(event__id=request.GET.get('e')).order_by('name')
                context['teams_sport'] = Team_sport.objects.filter(team__event__id=request.GET.get('e')).order_by('team','sport','-sexo')   
                context['event_sports'] = Event_sport.objects.filter(event__id=request.GET.get('e'))
        else:
            context['event'] = user.event_user
            context['teams'] = Team.objects.filter(unit=user.unit, event=user.event_user).order_by('name')
            context['teams_sport'] = Team_sport.objects.filter(team__unit=user.unit, team__event=user.event_user).order_by('team','sport','-sexo')   
            context['event_sports'] = Event_sport.objects.filter(event=user.event_user)
        return render(request, 'badge.html', context)
    else:
        print(request.POST)
        event = Event.objects.get(id=request.POST.get('event_data'))
        if 'team_sport_in' in request.POST: 
            print('team_sport_in')
            players = Player_team_sport.objects.filter(team_sport__id=request.POST.get('team_sport_in'), team_sport__event=event)
            team_sport_badge = Team_sport.objects.get(id=request.POST.get('team_sport_in'), event=event)
            if len(players) == 0:
                messages.error(request, "Não tem nenhum atleta cadastrado!")
                print('team_sport_in-zero')
            else:
                namebadge = f'{ team_sport_badge.sport.get_sport_display() }-{ team_sport_badge.team.name }-jifs'
                print('team_sport_in-criar')
                return generate_badges(players, '7',namebadge, event.id)
        elif 'team_in' in request.POST: 
            print('team_sport_in')
            players_qs = Player_team_sport.objects.filter(team_sport__team__id=request.POST.get('team_in'), team_sport__event=event)
            team = Team.objects.get(id=request.POST.get('team_in'))

            seen_players = set()
            players = []

            for pts in players_qs:
                if pts.player_id not in seen_players:
                    players.append(pts)
                    seen_players.add(pts.player_id)

            if len(players) == 0:
                messages.error(request, "Não tem nenhum atleta cadastrado!")
                print('team_in-zero')
            else:
                namebadge = f'{ team.name }-jifs'
                print('team_in-criar')
                return generate_badges(players, '7',namebadge, event.id)
        elif 'all_voluntary' in request.POST:
            print('all_voluntary')
            if user.is_staff: voluntary = Voluntary.objects.filter(type_voluntary=0, event=event)
            else: voluntary = Voluntary.objects.filter(type_voluntary=0, admin=user, event=event)
            if len(voluntary) == 0:
                messages.error(request, "Não tem nenhum voluntário cadastrado!")
                print('all_voluntary-zero')
            else:
                print("a: ",voluntary)
                namebadge = 'voluntarios-jifs'
                return generate_badges(voluntary, '0',namebadge, event.id)
        elif 'all_support' in request.POST:
            print('all_support-zero')
            if user.is_staff: voluntary = Voluntary.objects.filter(type_voluntary=2, event=event)
            else: voluntary = Voluntary.objects.filter(type_voluntary=2, admin=user, event=event)
            if len(voluntary) == 0:
                messages.error(request, "Não tem nenhum membro do apoio cadastrado!")
                print('all_support-zero')
            else:
                namebadge = 'apoio-jifs'
                return generate_badges(voluntary, '2',namebadge, event.id)
        elif 'all_organization' in request.POST:
            print('all_organization-zero')
            if user.is_staff: voluntary = Voluntary.objects.filter(type_voluntary=5, event=event)
            else: voluntary = Voluntary.objects.filter(type_voluntary=5, admin=user, event=event)
            if len(voluntary) == 0:
                messages.error(request, "Não tem nenhum membro da organização cadastrado!")
                print('all_organization-zero')
            else:
                namebadge = 'apoio-jifs'
                return generate_badges(voluntary, '5',namebadge, event.id)
        elif 'all_trainee' in request.POST:
            if user.is_staff: voluntary = Voluntary.objects.filter(type_voluntary=3, event=event)
            else: voluntary = Voluntary.objects.filter(type_voluntary=3, admin=user, event=event)
            if len(voluntary) == 0:
                messages.error(request, "Não tem nenhum estagiário cadastrado!")
            else:
                namebadge = 'estagiario-jifs'
                return generate_badges(voluntary, '3',namebadge, event.id)
        elif 'all_technician' in request.POST:
            if user.is_staff: voluntary = Voluntary.objects.filter(type_voluntary=1, event=event)
            else: voluntary = Voluntary.objects.filter(type_voluntary=1, admin=user, event=event)
            if len(voluntary) == 0:
                messages.error(request, "Não tem nenhum técnico cadastrado!")
            else:
                namebadge = 'tecnico-modalidade-jifs'
                return generate_badges(voluntary, '1',namebadge, event.id)
        elif 'all_head' in request.POST:
            if user.is_staff: voluntary = Voluntary.objects.filter(type_voluntary=4, event=event)
            else: voluntary = Voluntary.objects.filter(type_voluntary=4, admin=user, event=event)
            if len(voluntary) == 0:
                messages.error(request, "Não tem nenhum chefe de delegação cadastrado!")
            else:
                namebadge = 'chefe-delegacao-jifs'
                return generate_badges(voluntary, '4',namebadge, event.id)
        elif 'all_arbitrator' in request.POST:
            if user.is_staff: voluntary = Voluntary.objects.filter(type_voluntary=6, event=event)
            else: voluntary = Voluntary.objects.filter(type_voluntary=6, admin=user, event=event)
            if len(voluntary) == 0:
                messages.error(request, "Não tem nenhum árbitro cadastrado!")
            else:
                namebadge = 'chefe-delegacao-jifs'
                return generate_badges(voluntary, '6',namebadge, event.id)
    if current_get_params:
        print('com url-zero')
        return redirect(f"{reverse('badge')}?{current_get_params}")
    else:
        print('sem url-zero')
        return redirect('badge')


@login_required(login_url="login")
@terms_accept_required
@permission_required('app.add_certificate', raise_exception=True)
@permission_required('app.view_certificate', raise_exception=True)
def generator_certificate(request):
    try:
        user = User.objects.get(id=request.user.id)
        if request.user.is_staff:
            team_sport = Team_sport.objects.all()
        else:
            team_sport = Team_sport.objects.filter(admin__id=request.user.id).order_by('team','sport','-sexo')
        certificate = Certificate.objects.filter(user=request.user.id)
        sport = Sport_types.choices
        if request.method == "GET":
            context = {
                'team_sport': team_sport,
                'sport': sport,
                'certificate': certificate,
                
            }
            return render(request, 'certificate.html', context)
        else:
            if 'certificate_delete' in request.POST:
                certificate_delete = request.POST.get('certificate_delete')
                certificate = Certificate.objects.get(id=certificate_delete)
                certificate.file.delete()
                certificate.delete()
                return redirect('certificate')
            elif 'certificate_all_delete' in request.POST:
                certificate = Certificate.objects.all()
                for i in certificate:
                    i.file.delete()
                    i.delete()
                return redirect('certificate')
            elif 'team-certificate' in request.POST:
                team_certificate = request.POST.get('team-certificate')
                if team_certificate.isdigit(): 
                    team_sport = get_object_or_404(Team_sport, id=team_certificate) 
                    players = Player_team_sport.objects.filter(team_sport=team_sport)
                    generate_certificates(players, user, '2')
                else:
                    if team_certificate == 'all_player':
                        players = Player_team_sport.objects.all()
                        generate_certificates(players, user, 0)
                    elif team_certificate == 'all_voluntary':
                        voluntary = Voluntary.objects.filter(type_voluntary=0, admin=user)
                        print("a: ",voluntary)
                        generate_certificates(voluntary, user, 1)
                    elif team_certificate == 'all_organization':
                        voluntary = Voluntary.objects.filter(type_voluntary=1, admin=user)
                        generate_certificates(voluntary, user, 1)
                    elif team_certificate == 'all_technician':
                        voluntary = Voluntary.objects.filter(type_voluntary=2, admin=user)
                        generate_certificates(voluntary, user, 1)
                    else:
                        for choice in Sport_types.choices:
                            if choice[1] == team_certificate:
                                sport_value = choice[0]
                                break
                        players = Player_team_sport.objects.filter(team_sport__sport=sport_value)
                return redirect('certificate')
            return redirect('certificate')
    except Exception as e:
        messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
    return redirect('certificate')

# =============================================================================
# Substitua APENAS a função generator_data no views.py
# =============================================================================

@login_required(login_url="login")
@terms_accept_required
def generator_data(request):
    user = User.objects.get(id=request.user.id)
    current_get_params = request.GET.urlencode()

    # ── GET ───────────────────────────────────────────────────────────────────
    if request.method == "GET":
        if request.user.is_staff:
            team_sport = Team_sport.objects.filter(
                players__isnull=False
            ).distinct().order_by('sport', '-sexo')
        elif request.user.type == 1:
            team_sport = Team_sport.objects.filter(
                players__isnull=False, event=user.event_user
            ).order_by('sport', '-sexo')
        else:
            team_sport = Team_sport.objects.filter(
                players__isnull=False, team__unit=user.unit
            ).order_by('sport', '-sexo')

        context = {
            'sexo': Sexo_types.choices,
            'team_sport': team_sport,
            'sports_general': Sport_types.choices,
            'events': Event.objects.all(),
        }

        # ── popula event / teams / event_sports por tipo de usuário ──────────
        if user.type == 0:
            # admin: precisa selecionar o evento via ?e=
            if 'e' in request.GET and request.GET.get('e') != '':
                event = Event.objects.get(id=request.GET.get('e'))
                context['event']        = event
                context['teams']        = Team.objects.filter(event=event).order_by('name')
                context['event_sports'] = Event_sport.objects.filter(event=event)

        elif user.type == 1:
            # coordenador: sempre vê o próprio evento, sem precisar de ?e=
            event = user.event_user
            if event:
                context['event']        = event
                context['teams']        = Team.objects.filter(event=event).order_by('name')
                context['event_sports'] = Event_sport.objects.filter(event=event)

        else:
            # usuário comum (type=2): vê o próprio evento, sem botões de campus
            if user.event_user:
                context['event']        = user.event_user
                context['event_sports'] = Event_sport.objects.filter(event=user.event_user)
            # não passa 'teams': tipo 2 não tem botão de campus individual

        return render(request, 'data.html', context)

    # ── POST ──────────────────────────────────────────────────────────────────
    cont = {
        'now':       timezone.now(),
        'user':      user,
        'logo_morea': request.build_absolute_uri('/static/images/logo_atum.png'),
    }
    status = False

    # resolve evento do POST
    event_id = request.POST.get('event_data')
    if not event_id or event_id == '0':
        messages.error(request, "Evento não identificado.")
        return redirect('data')

    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        messages.error(request, "Evento não encontrado.")
        return redirect('data')

    # ── validação de acesso por tipo ─────────────────────────────────────────
    if user.type in (1, 2) and event != user.event_user:
        messages.error(request, "Você não tem permissão para acessar este evento.")
        return redirect('data')

    # ── cabeçalho do PDF (logo do evento) ────────────────────────────────────
    if event.logo:
        cont['logo_ifs'] = request.build_absolute_uri(event.logo.url)
    else:
        cont['logo_ifs'] = request.build_absolute_uri('/static/images/logo-jiifs-2025.jpg')
    cont['event'] = event

    logo_ifs = cont['logo_ifs']
    img = ImageReader(logo_ifs)
    largura, altura = img.getSize()
    if largura == altura:
        cont['logo_event_type'] = 0
    elif largura > altura:
        cont['logo_event_type'] = 1
    else:
        cont['logo_event_type'] = 2

    # ─────────────────────────────────────────────────────────────────────────
    # BLOCO: dados gerais (admin + coordenador)
    # ─────────────────────────────────────────────────────────────────────────
    if 'all_data' in request.POST:
        if user.type not in (0, 1) and not user.is_staff:
            messages.error(request, "Sem permissão para este relatório.")
            return redirect('data')

        name_html = 'data-general'
        name_pdf  = 'dados_gerais'

        qnt_players      = Player.objects.filter(event=event).count()
        qnt_players_fem  = Player.objects.filter(sexo=1, event=event).count()
        qnt_players_masc = Player.objects.filter(sexo=0, event=event).count()

        cont['qnt_players']      = qnt_players
        cont['qnt_players_fem']  = qnt_players_fem
        cont['qnt_players_masc'] = qnt_players_masc
        cont['qnt_teams']        = Team_sport.objects.filter(event=event).count()
        cont['qnt_voluntary_0']  = Voluntary.objects.filter(type_voluntary=0, event=event).count()
        cont['qnt_voluntary_1']  = Voluntary.objects.filter(type_voluntary=1, event=event).count()
        cont['qnt_voluntary_2']  = Voluntary.objects.filter(type_voluntary=2, event=event).count()
        cont['qnt_voluntary_3']  = Voluntary.objects.filter(type_voluntary=3, event=event).count()
        cont['qnt_voluntary_4']  = Voluntary.objects.filter(type_voluntary=4, event=event).count()

        # guarda divisão por zero
        if qnt_players > 0:
            cont['porcent_fem']  = (qnt_players_fem  * 100) / qnt_players
            cont['porcent_masc'] = (qnt_players_masc * 100) / qnt_players
        else:
            cont['porcent_fem']  = 0
            cont['porcent_masc'] = 0

    # ─────────────────────────────────────────────────────────────────────────
    # BLOCO: inscrições (admin + coordenador)
    # ─────────────────────────────────────────────────────────────────────────
    elif 'enrollment' in request.POST:
        if user.type not in (0, 1) and not user.is_staff:
            messages.error(request, "Sem permissão para este relatório.")
            return redirect('data')

        name_html = 'data-base-enrollment'
        name_pdf  = 'relatório_de_inscrições'
        teams = Team_sport.objects.prefetch_related('players').filter(
            team__event=event
        ).order_by('team', 'sport', '-sexo')
        if not teams.exists():
            messages.error(request, "Não há equipes em modalidades ou atletas cadastrados.")
            status = True
        cont['teams'] = teams

    # ─────────────────────────────────────────────────────────────────────────
    # BLOCO: dados dos times (todos por campus)
    # ─────────────────────────────────────────────────────────────────────────
    elif 'all_team' in request.POST:
        name_html = 'data-base-campus'
        name_pdf  = 'dados_campus'

        if user.is_staff or user.type in (0, 1):
            teams = Team_sport.objects.prefetch_related('players').filter(
                event=event
            ).order_by('sport', '-sexo')
        else:
            teams = Team_sport.objects.prefetch_related('players').filter(
                team__unit=user.unit, event=event
            ).order_by('sport', '-sexo')

        if not teams.exists():
            messages.error(request, "Não há equipes ou atletas cadastrados.")
            status = True
        cont['teams'] = teams
        cont['infor'] = "campus x modalidade x atletas"

    # ─────────────────────────────────────────────────────────────────────────
    # BLOCO: partidas (admin + coordenador)
    # ─────────────────────────────────────────────────────────────────────────
    elif 'all_match' in request.POST:
        if user.type not in (0, 1) and not user.is_staff:
            messages.error(request, "Sem permissão para este relatório.")
            return redirect('data')

        name_html = 'data-base-match'
        name_pdf  = 'partidas'
        matchs = Match.objects.filter(event=event).prefetch_related('teams__team')
        sport  = Sport_types.choices
        context_matchs = [
            {
                'match': match,
                'sport': sport,
                'times': list(match.teams.all()),
            }
            for match in matchs
        ]
        if not matchs.exists():
            messages.error(request, "Não há nenhuma partida programada.")
            status = True
        cont['context'] = context_matchs

    # ─────────────────────────────────────────────────────────────────────────
    # BLOCO: comissão técnica
    # ─────────────────────────────────────────────────────────────────────────
    elif 'all_eqp' in request.POST:
        name_html = 'data-base-eqp'
        name_pdf  = 'dados_equipe'
        cont['infor'] = f"comissão técnica do {event.name}"

        if user.is_staff or user.type in (0, 1):
            voluntary = Voluntary.objects.filter(event=event).order_by('-type_voluntary')
        else:
            voluntary = Voluntary.objects.filter(admin=user, event=event).order_by('-type_voluntary')

        if not voluntary.exists():
            messages.error(request, "Não há membros da comissão cadastrados.")
            status = True
        cont['team'] = voluntary

    # ─────────────────────────────────────────────────────────────────────────
    # BLOCO: todos os atletas
    # ─────────────────────────────────────────────────────────────────────────
    elif 'all_players' in request.POST:
        name_html = 'data-base'
        name_pdf  = 'dados_atletas'
        cont['infor'] = "atletas"

        if user.is_staff or user.type in (0, 1):
            players = Player.objects.filter(event=event).order_by('-sexo')
        else:
            players = Player.objects.filter(unit=user.unit, event=event).order_by('-sexo')

        if not players.exists():
            messages.error(request, "Não há atletas cadastrados.")
            status = True
        cont['players'] = players

    # ─────────────────────────────────────────────────────────────────────────
    # BLOCO: atletas femininas
    # ─────────────────────────────────────────────────────────────────────────
    elif 'all_players_fem' in request.POST:
        name_html = 'data-base'
        name_pdf  = 'dados_atletas_femininas'

        if user.is_staff or user.type in (0, 1):
            players = Player.objects.filter(sexo=1, event=event).order_by('name')
        else:
            players = Player.objects.filter(sexo=1, unit=user.unit, event=event).order_by('name')

        if not players.exists():
            messages.error(request, "Não há atletas do sexo feminino cadastradas.")
            status = True
        cont['players'] = players
        cont['infor']   = "atletas do sexo feminino"
        cont['type']    = True

    # ─────────────────────────────────────────────────────────────────────────
    # BLOCO: atletas masculinos
    # ─────────────────────────────────────────────────────────────────────────
    elif 'all_players_masc' in request.POST:
        name_html = 'data-base'
        name_pdf  = 'dados_atletas_masculinos'

        if user.is_staff or user.type in (0, 1):
            players = Player.objects.filter(sexo=0, event=event).order_by('name')
        else:
            players = Player.objects.filter(sexo=0, unit=user.unit, event=event).order_by('name')

        if not players.exists():
            messages.error(request, "Não há atletas do sexo masculino cadastrados.")
            status = True
        cont['players'] = players
        cont['infor']   = "atletas do sexo masculino"
        cont['type']    = True

    # ─────────────────────────────────────────────────────────────────────────
    # BLOCO: campus individual (admin + coordenador)
    # ─────────────────────────────────────────────────────────────────────────
    elif 'team_in' in request.POST:
        if user.type not in (0, 1) and not user.is_staff:
            messages.error(request, "Sem permissão para este relatório.")
            return redirect('data')

        team_id = request.POST.get('team_in')
        try:
            team = Team.objects.get(id=team_id, event=event)
        except Team.DoesNotExist:
            messages.error(request, "Campus não encontrado neste evento.")
            return redirect('data')

        name_html = 'data-base-campus-individual'
        name_pdf  = f'atletas_{team.name}'

        players = (
            Player_team_sport.objects
            .filter(team_sport__team=team, player__event=event)
            .select_related('player', 'team_sport__team', 'team_sport__sport')
            .order_by('player__name')
        )

        if not players.exists():
            messages.error(request, "Não há atletas cadastrados neste campus.")
            status = True

        cont['team']   = team
        cont['players'] = players
        cont['infor']   = "atletas"
        cont['campus']  = team.name

    # ─────────────────────────────────────────────────────────────────────────
    # BLOCO: por modalidade
    # ─────────────────────────────────────────────────────────────────────────
    elif 'all_players_sport' in request.POST:
        data = request.POST.get('all_players_sport')
        event_sport = get_object_or_404(Event_sport, id=data, event=event)

        name_html = 'data-base-campus'
        name_pdf  = f'atletas_{event_sport.get_sport_display()}'

        if user.is_staff or user.type in (0, 1):
            teams = (
                Team_sport.objects
                .filter(sport=event_sport, event=event)
                .prefetch_related('players')
                .order_by('-sport', '-sexo')
            )
        else:
            # usuário comum: apenas times da sua unidade
            teams = (
                Team_sport.objects
                .filter(sport=event_sport, event=event, team__unit=user.unit)
                .prefetch_related('players')
                .order_by('-sport', '-sexo')
            )

        if not teams.exists():
            messages.error(request, "Não há equipes ou atletas cadastrados nessa modalidade.")
            status = True
        cont['teams'] = teams

    else:
        messages.error(request, "Nenhum relatório foi selecionado.")
        return redirect('data')

    # ── gera o PDF ────────────────────────────────────────────────────────────
    if status:
        if current_get_params:
            return redirect(f"{reverse('data')}?{current_get_params}")
        return redirect('data')

    html_string = render_to_string(f'generator/{name_html}.html', cont)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{name_pdf}.pdf"'
    HTML(string=html_string).write_pdf(response)

    return response

def generate_spreadsheet(players, event_name, label):
    wb = Workbook()

    thin = Side(style='thin', color='000000')
    medium = Side(style='medium', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font = Font(name='Arial', size=10, bold=True)
    body_font = Font(name='Arial', size=10)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    fill_header_pag = PatternFill(fill_type='solid', fgColor='BDBDBD')
    fill_header_end = PatternFill(fill_type='solid', fgColor='AEABAB')
    fill_white = PatternFill(fill_type='solid', fgColor='FFFFFF')

    def make_header(ws, headers, widths, fill_color):
        ws.row_dimensions[1].height = 35

        for i, (header, width) in enumerate(zip(headers, widths), start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
            cell = ws.cell(row=1, column=i, value=header)
            cell.font = header_font
            cell.fill = PatternFill(fill_type='solid', fgColor=fill_color)
            cell.alignment = center
            cell.border = Border(left=thin, right=thin, top=medium, bottom=thin)

    def style_row(ws, row_num, num_cols, center_cols=()):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = body_font
            cell.fill = fill_white
            cell.border = border
            cell.alignment = center if col in center_cols else left

    # ABA 1 - PAGAMENTO
    ws1 = wb.active
    ws1.title = 'PLANILHA DE PAGAMENTO'

    headers_pag = [
        'SEQ',
        'MATRÍCULA',
        'NOME DO ALUNO',
        'CPF/CHAVE PIX [Somente Números]',
        'VALOR',
        'CURSO',
        'AUXÍLIO [Descrição]',
        'EVENTO',
        'VALOR TOTAL',
    ]

    widths_pag = [5.86, 13.29, 44.14, 24.00, 14.29, 29.71, 20.00, 20.00, 16.00]
    make_header(ws1, headers_pag, widths_pag, 'BDBDBD')

    for seq, player in enumerate(players, start=1):
        ws1.append([
            seq,
            player.registration or '',
            player.name or '',
            (player.cpf or '').replace('.', '').replace('-', '').replace('/', ''),
            '',          # VALOR
            player.course or '',
            '',          # AUXÍLIO [Descrição]
            event_name,  # EVENTO
            '',          # VALOR TOTAL fica vazio nas linhas
        ])

        style_row(ws1, ws1.max_row, len(headers_pag), center_cols=(1, 2, 4, 5, 8, 9))

    # linha de total geral
    first_data_row = 2
    last_data_row = ws1.max_row
    total_row = last_data_row + 1

    # deixa a linha inteira estilizada
    style_row(ws1, total_row, len(headers_pag), center_cols=(1, 2, 4, 5, 8, 9))

    # rótulo do total
    ws1.cell(row=total_row, column=8, value='TOTAL GERAL')
    ws1.cell(row=total_row, column=8).font = header_font
    ws1.cell(row=total_row, column=8).alignment = center
    ws1.cell(row=total_row, column=8).border = border
    ws1.cell(row=total_row, column=8).fill = fill_white

    # fórmula somando toda a coluna VALOR
    if last_data_row >= first_data_row:
        ws1.cell(row=total_row, column=9, value=f'=SUM(E{first_data_row}:E{last_data_row})')
    else:
        ws1.cell(row=total_row, column=9, value='')

    ws1.cell(row=total_row, column=9).font = header_font
    ws1.cell(row=total_row, column=9).alignment = center
    ws1.cell(row=total_row, column=9).border = border
    ws1.cell(row=total_row, column=9).fill = fill_white

    for row in range(2, ws1.max_row + 1):
        ws1.cell(row=row, column=5).number_format = 'R$ #,##0.00'
        ws1.cell(row=row, column=9).number_format = 'R$ #,##0.00'

    # ABA 2 - ENDEREÇO
    ws2 = wb.create_sheet('PLANILHA DE ENDEREÇO')

    headers_end = [
        'SEQ.',
        'MATRÍCULA',
        'NOME DO ALUNO',
        'CPF / CHAVE PIX [Somente Números]',
        'CURSO',
        'ENDEREÇO COMPLETO [Rua, Bairro e Número]',
        'CEP [Somente Números]',
        'MUNICÍPIO/UF',
    ]

    widths_end = [5.86, 13.29, 37.57, 24.00, 29.71, 63.86, 18.00, 26.0]
    make_header(ws2, headers_end, widths_end, 'AEABAB')

    for seq, player in enumerate(players, start=1):
        ws2.append([
            seq,
            player.registration or '',
            player.name or '',
            (player.cpf or '').replace('.', '').replace('-', '').replace('/', ''),
            player.course or '',
            player.address or '',
            (player.cep or '').replace('-', '').replace('.', ''),
            player.municipality or '',
        ])

        style_row(ws2, ws2.max_row, len(headers_end), center_cols=(1, 2, 4, 7, 8))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_label = (
        label.replace(' ', '_')
        .replace('/', '_')
        .replace('\\', '_')
        .lower()
    )

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="planilha_{safe_label}.xlsx"'
    return response


@login_required(login_url="login")
@terms_accept_required
def generator_spreadsheet(request):
    user = request.user

    # permite:
    # - quem tem a permissão app.data
    # - técnico (type == 2)
    if not (user.has_perm('app.data') or user.type in [1, 2]):
        raise PermissionDenied

    current_get_params = request.GET.urlencode()

    def _redirect():
        if current_get_params:
            return redirect(f"{reverse('spreadsheet')}?{current_get_params}")
        return redirect('spreadsheet')

    def _dedup_from_pts(pts_qs):
        seen = set()
        result = []
        for pts in pts_qs:
            if pts.player_id not in seen:
                result.append(pts.player)
                seen.add(pts.player_id)
        return result

    # GET — monta contexto
    if request.method == "GET":
        context = {}

        if user.type == 0:
            context['events'] = Event.objects.all()

            selected_event = request.GET.get('e')
            if selected_event:
                try:
                    event = Event.objects.get(id=selected_event)
                    context['event'] = event
                    context['select_event'] = selected_event
                    context['teams'] = Team.objects.filter(event=event).order_by('name')
                    context['event_sports'] = Event_sport.objects.filter(event=event)
                except Event.DoesNotExist:
                    messages.error(request, "Evento selecionado não foi encontrado.")

        elif user.type == 1:
            event = user.event_user
            if event:
                context['event'] = event
                context['teams'] = Team.objects.filter(event=event).order_by('name')
                context['event_sports'] = Event_sport.objects.filter(event=event)
            else:
                messages.error(request, "Seu usuário não possui evento vinculado.")

        elif user.type == 2:
            if user.event_user and user.team:
                event = user.event_user
                context['event'] = event
                context['my_team'] = user.team
                context['my_team_sports'] = (
                    Team_sport.objects
                    .filter(team=user.team, event=event)
                    .select_related('sport')
                    .order_by('sport__sport', 'sexo')
                )
            else:
                messages.error(request, "Seu usuário técnico precisa ter evento e campus vinculados.")

        return render(request, 'spreadsheet.html', context)

    # POST — gera o xlsx
    event_id = request.POST.get('event_data')
    if not event_id:
        messages.error(request, "Evento não identificado.")
        return _redirect()

    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        messages.error(request, "Evento não encontrado.")
        return _redirect()

    # segurança extra por tipo de usuário
    if user.type in [1, 2] and user.event_user != event:
        messages.error(request, "Você não tem permissão para acessar este evento.")
        return _redirect()

    players = None
    label = event.name

    # GERAL
    if 'p_todos' in request.POST:
        if user.type == 2:
            players = list(
                Player.objects.filter(event=event, admin=user).order_by('name')
            )
        else:
            players = list(
                Player.objects.filter(event=event).order_by('name')
            )
        label = f'todos_{event.name}'

    elif 'p_masc' in request.POST:
        if user.type == 2:
            players = list(
                Player.objects.filter(event=event, admin=user, sexo=0).order_by('name')
            )
        else:
            players = list(
                Player.objects.filter(event=event, sexo=0).order_by('name')
            )
        label = f'masculino_{event.name}'

    elif 'p_fem' in request.POST:
        if user.type == 2:
            players = list(
                Player.objects.filter(event=event, admin=user, sexo=1).order_by('name')
            )
        else:
            players = list(
                Player.objects.filter(event=event, sexo=1).order_by('name')
            )
        label = f'feminino_{event.name}'

    # POR CAMPUS
    elif 'p_team' in request.POST:
        if user.type not in [0, 1]:
            messages.error(request, "Você não tem permissão para gerar esta planilha.")
            return _redirect()

        try:
            team = Team.objects.get(id=request.POST.get('p_team'), event=event)
        except Team.DoesNotExist:
            messages.error(request, "Campus não encontrado neste evento.")
            return _redirect()

        pts_qs = (
            Player_team_sport.objects
            .filter(team_sport__team=team, player__event=event)
            .select_related('player')
            .order_by('player__name')
        )
        players = _dedup_from_pts(pts_qs)
        label = f'{team.name}_{event.name}'

    # POR MODALIDADE — superusuário/coordenador
    elif 'p_sport' in request.POST:
        if user.type not in [0, 1]:
            messages.error(request, "Você não tem permissão para gerar esta planilha.")
            return _redirect()

        try:
            event_sport = Event_sport.objects.get(id=request.POST.get('p_sport'), event=event)
        except Event_sport.DoesNotExist:
            messages.error(request, "Modalidade não encontrada neste evento.")
            return _redirect()

        pts_qs = (
            Player_team_sport.objects
            .filter(team_sport__sport=event_sport, player__event=event)
            .select_related('player')
            .order_by('player__name')
        )
        players = _dedup_from_pts(pts_qs)
        label = f'{event_sport.get_sport_display()}_{event.name}'

    # POR MODALIDADE + GÊNERO — técnico
    elif 'p_team_sport' in request.POST:
        try:
            team_sport_obj = Team_sport.objects.select_related('team', 'sport').get(
                id=request.POST.get('p_team_sport'),
                event=event
            )
        except Team_sport.DoesNotExist:
            messages.error(request, "Modalidade da equipe não encontrada.")
            return _redirect()

        if user.type == 2 and team_sport_obj.team != user.team:
            messages.error(request, "Você não tem permissão para gerar esta planilha.")
            return _redirect()

        pts_qs = (
            Player_team_sport.objects
            .filter(team_sport=team_sport_obj)
            .select_related('player')
            .order_by('player__name')
        )
        players = _dedup_from_pts(pts_qs)
        label = (
            f'{team_sport_obj.sport.get_sport_display()}_'
            f'{team_sport_obj.get_sexo_display()}_'
            f'{event.name}'
        )

    else:
        messages.error(request, "Nenhum filtro de planilha foi enviado.")
        return _redirect()

    if not players:
        messages.error(request, "Não há atletas cadastrados para os filtros selecionados.")
        return _redirect()

    return generate_spreadsheet(players, event.name, label)


# Dashboard de acesso


def _has_user_document(user):
    doc = getattr(user, "document", None)
    return bool(doc and str(doc).strip())


def _format_dt(dt):
    if not dt:
        return ""
    if timezone.is_aware(dt):
        dt = timezone.localtime(dt)
    return dt.strftime("%d/%m/%Y %H:%M")


def _get_user_access_qs(user, selected_date=None, event=None):
    qs = AccessLog.objects.filter(user=user)

    if event:
        qs = qs.filter(event=event)

    if selected_date:
        qs = qs.filter(accessed_at__date=selected_date)

    return qs.order_by("accessed_at")


def _build_user_dashboard_row(user, selected_date=None, event=None):
    access_qs = _get_user_access_qs(user, selected_date=selected_date, event=event)

    total_acessos = access_qs.count()
    primeiro_acesso = access_qs.first()
    ultimo_acesso = access_qs.order_by("-accessed_at").first()

    return {
        "user": user,
        "primeiro_login": primeiro_acesso.accessed_at if primeiro_acesso else None,
        "ultimo_login": ultimo_acesso.accessed_at if ultimo_acesso else None,
        "total_acessos": total_acessos,
        "accepted": bool(getattr(user, "accepted", False)),
        "accepted_at": getattr(user, "accepted_at", None),
        "has_document": _has_user_document(user),
    }


@permission_required("app.view_dashboard_acesso", raise_exception=True)
@login_required(login_url="login")
@terms_accept_required
def dashboard_acesso(request):
    user = request.user

    if not user.is_superuser and getattr(user, "type", None) not in [0, 1]:
        messages.error(request, "Você não tem permissão para acessar essa página.")
        return redirect("Home")

    user_model = get_user_model()

    selected_event_id = request.GET.get("e", "").strip()
    selected_user_id = request.GET.get("u", "").strip()
    selected_date_raw = request.GET.get("d", "").strip()

    try:
        selected_date_obj = date.fromisoformat(selected_date_raw) if selected_date_raw else None
    except ValueError:
        selected_date_obj = None
        selected_date_raw = ""

    events = None
    event = None

    if user.type == 0:
        events = Event.objects.all().order_by("name")

        if selected_event_id:
            try:
                event = Event.objects.get(id=selected_event_id)
            except Event.DoesNotExist:
                messages.error(request, "Evento selecionado não foi encontrado.")
                selected_event_id = ""
                event = None

    elif user.type == 1:
        event = user.event_user
        if not event:
            messages.error(request, "Seu usuário não possui evento vinculado.")
            return redirect("Home")
        selected_event_id = str(event.id)

    if user.type == 0 and not event:
        context = {
            "events": events,
            "event": None,
            "select_event": selected_event_id,
            "all_users": [],
            "selected_user_id": "",
            "selected_user_obj": None,
            "selected_date": selected_date_raw,
            "total_users": 0,
            "total_accepted": 0,
            "total_docs": 0,
            "total_pendentes": 0,
            "acessos_hoje": 0,
            "acessos_7dias": [],
            "user_rows": [],
        }
        return render(request, "dashboard_acesso.html", context)

    all_users = (
        user_model.objects
        .exclude(is_superuser=True)
        .filter(event_user=event)
        .order_by("username")
    )

    selected_user_obj = None
    users_qs = all_users

    if selected_user_id:
        users_qs = all_users.filter(id=selected_user_id)
        selected_user_obj = users_qs.first()

    hoje = timezone.localdate()

    total_users = all_users.count()
    total_accepted = all_users.filter(accepted=True).count()
    total_docs = sum(1 for u in all_users if _has_user_document(u))
    total_pendentes = sum(
        1
        for u in all_users
        if (not getattr(u, "accepted", False)) or (not _has_user_document(u))
    )

    acessos_hoje = (
        AccessLog.objects.filter(
            event=event,
            accessed_at__date=hoje
        )
        .values("user")
        .distinct()
        .count()
    )

    acessos_7dias_raw = []
    max_count = 0

    for i in range(6, -1, -1):
        dia = hoje - timedelta(days=i)
        count = (
            AccessLog.objects.filter(
                event=event,
                accessed_at__date=dia
            )
            .values("user")
            .distinct()
            .count()
        )
        acessos_7dias_raw.append({"dia": dia, "count": count})
        max_count = max(max_count, count)

    acessos_7dias = []
    for item in acessos_7dias_raw:
        count = item["count"]
        pct = max(6, round((count / max_count) * 100)) if max_count > 0 else 0

        acessos_7dias.append(
            {
                "label": item["dia"].strftime("%d/%m"),
                "count": count,
                "pct": pct,
            }
        )

    user_rows = []
    for u in users_qs:
        row = _build_user_dashboard_row(
            u,
            selected_date=selected_date_obj,
            event=event,
        )
        user_rows.append(row)

    context = {
        "events": events,
        "event": event,
        "select_event": selected_event_id,
        "all_users": all_users,
        "selected_user_id": selected_user_id,
        "selected_user_obj": selected_user_obj,
        "selected_date": selected_date_raw,
        "total_users": total_users,
        "total_accepted": total_accepted,
        "total_docs": total_docs,
        "total_pendentes": total_pendentes,
        "acessos_hoje": acessos_hoje,
        "acessos_7dias": acessos_7dias,
        "user_rows": user_rows,
    }
    return render(request, "dashboard_acesso.html", context)


@login_required(login_url="login")
@terms_accept_required
def dashboard_acesso_user_detail(request, user_id):
    user = request.user

    if not user.is_superuser and getattr(user, "type", None) not in [0, 1]:
        return JsonResponse({"error": "Sem permissão"}, status=403)

    selected_event_id = request.GET.get("e", "").strip()

    event = None
    if user.type == 0:
        if selected_event_id:
            event = Event.objects.filter(id=selected_event_id).first()
    elif user.type == 1:
        event = user.event_user

    if not event:
        return JsonResponse({"error": "Evento não selecionado"}, status=400)

    user_model = get_user_model()
    u = user_model.objects.filter(
        id=user_id,
        event_user=event
    ).first()

    if not u:
        return JsonResponse({"error": "Não encontrado"}, status=404)

    row_data = _build_user_dashboard_row(u, event=event)

    access_logs_qs = (
        AccessLog.objects.filter(user=u, event=event)
        .select_related("session", "event")
        .order_by("-accessed_at")[:10]
    )

    service_map = {
        "team_manage": "Times",
        "player_manage": "Atletas",
        "voluntary_manage": "Equipe",
        "games": "Partidas",
        "badge": "Crachás",
        "data": "Relatórios",
        "spreadsheet": "Planilhas",
        "certificate": "Certificados",
        "attachments": "Anexos",
        "scoreboard": "Placar",
        "user_manage": "Usuários",
        "Home": "Início",
        "event_manage": "Eventos",
    }

    sessions_data = []

    for log in access_logs_qs:
        session_data = {}
        services = []

        try:
            if log.session:
                session_data = log.session.get_decoded() or {}
                visited = session_data.get("visited_urls", []) or []
                services = list(
                    dict.fromkeys(
                        [service_map.get(url, url) for url in visited if url]
                    )
                )
        except Exception:
            session_data = {}
            services = []

        sessions_data.append(
            {
                "date": _format_dt(log.accessed_at),
                "services": services,
                "ip": log.ip_address or session_data.get("client_ip", ""),
            }
        )

    full_name = (
        getattr(u, "get_full_name", lambda: "")().strip()
        or getattr(u, "first_name", "")
        or u.username
    )

    data = {
        "username": u.username,
        "full_name": full_name,
        "first_name": u.first_name or "",
        "email": u.email or "",
        "telefone": getattr(u, "telefone", "") or "",
        "type_display": u.get_type_display() if hasattr(u, "get_type_display") else "",
        "photo": u.photo.url if getattr(u, "photo", None) else "",
        "event_user": u.event_user.name if getattr(u, "event_user", None) else "",
        "unit": u.unit.name if getattr(u, "unit", None) else "",
        "team": u.team.name if getattr(u, "team", None) else "",
        "accepted": row_data["accepted"],
        "accepted_at": _format_dt(row_data["accepted_at"]),
        "has_document": row_data["has_document"],
        "document_url": u.document.url if _has_user_document(u) else "",
        "primeiro_login": _format_dt(row_data["primeiro_login"]),
        "ultimo_login": _format_dt(row_data["ultimo_login"]),
        "total_acessos": row_data["total_acessos"],
        "sessions": sessions_data,
    }

    return JsonResponse(data)

def verificar_foto(url_name):
    print("url: ", url_name)
    list = ['person.png','team.png']
    url = url_name.split('/')
    delete_photo = url[len(url) - 1]
    if delete_photo in list:
        return False
    return True

def type_file(request, rest, file, text):
    ext = os.path.splitext(file.name)[1].lower().strip()
    if ext not in rest:
        messages.error(request, text)
        return True
    return False

