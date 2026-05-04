"""
Views extraídas automaticamente do views.py monolítico.
Módulo: generators.py
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse, QueryDict
from app.models import Sexo_types, Settings_access, UserSession, Event_unit, Event_badge, Detailed, Status, Authenticity, Match_referee, Type_referee, Replacement, Group_phase, Phase, Phase_types, Campus_types, Help, Type_penalties, Detailed, Activity, Statement, Point_types, Event, Event_sport, Statement_user, Users_types, Type_service, Certificate, Attachments, Volley_match, Player, Sport_types, Voluntary, Penalties, Occurrence, Time_pause, Team, Point, Team_sport, Player_team_sport, Match, Team_match, Player_match, Assistance,  Banner, Terms_Use
from django.db.models import Count, Q, Prefetch
from app.decorators import time_restriction
from django.contrib import messages
from django.db import IntegrityError
from django.templatetags.static import static
from django.contrib.auth.models import User, Group, Permission
from django.contrib.sessions.models import Session
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth import login as auth_login, authenticate, logout, get_user_model
from django.template.loader import render_to_string
from app.forms import Terms_UseForm
from datetime import date, datetime
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from app.generators import generate_certificates, generate_badges, generate_events, generate_timer
import time, pytz, os, random
from django.core.files.base import ContentFile
from weasyprint import HTML
from django.utils import timezone
from app.decorators import terms_accept_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.urls import reverse
from django.contrib.auth import update_session_auth_hash
from django.core.exceptions import PermissionDenied
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import pytz
from datetime import datetime as dt
from collections import Counter
from django.db.models.functions import TruncDate
from django.db.models import Count
from datetime import date, timedelta
from datetime import date, timedelta
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, permission_required
from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from app.models import AccessLog, UserSession
from app.decorators import terms_accept_required
from datetime import date, timedelta
from app.models import AccessLog, Event

# Imports locais do projeto
from app.helpers import (
    acesso_evento, acesso_team, acesso_team_sport,
    acesso_player, acesso_match, verificar_foto,
    type_file, calcular_idade, generate_authenticity,
    check_gender_compatibility, player_queryset_for_team,
    SEXO_NAMES,
)
from app.services.pdf import build_pdf_context, generate_pdf_response, generate_qr_base64
from app.services.points import calculate_points, calculate_penalties_count, get_aces_count
from app.services.volleyball import get_ordered_team_matches, get_ordered_sets
from app.services.password import generate_random_password

@login_required(login_url="login")
@terms_accept_required
def generator_badge(request):
    current_get_params = request.GET.urlencode()
    user = User.objects.get(id=request.user.id)

    def redirect_badge():
        if current_get_params:
            return redirect(f"{reverse('badge')}?{current_get_params}")
        return redirect('badge')

    def voluntary_queryset(event, type_voluntary=None):
        qs = Voluntary.objects.filter(event=event)

        if type_voluntary is not None:
            qs = qs.filter(type_voluntary=type_voluntary)

        # Admin: qualquer unidade do evento selecionado
        if user.is_staff or user.type == 0:
            return qs.order_by('type_voluntary', 'unit__name', 'name')

        # Coordenador de evento: todas as unidades do próprio evento
        if user.type == 1:
            return qs.filter(
                event=user.event_user
            ).order_by('type_voluntary', 'unit__name', 'name')

        # Usuário comum / chefe de delegação: apenas própria unidade
        return qs.filter(
            event=user.event_user,
            unit=user.unit
        ).order_by('type_voluntary', 'unit__name', 'name')

    if request.method == "GET":
        context = {}

        if user.is_staff or user.type == 0:
            context['events'] = Event.objects.all()

            if 'e' in request.GET and request.GET.get('e') != '':
                event = Event.objects.get(id=request.GET.get('e'))
                context['event'] = event
                context['teams'] = Team.objects.filter(event=event).order_by('name')
                context['teams_sport'] = Team_sport.objects.filter(
                    event=event
                ).order_by('team', 'sport', '-sexo')
                context['event_sports'] = Event_sport.objects.filter(event=event)

        else:
            event = user.event_user
            context['event'] = event

            if user.type == 1:
                context['teams'] = Team.objects.filter(
                    event=event
                ).order_by('name')

                context['teams_sport'] = Team_sport.objects.filter(
                    event=event
                ).order_by('team', 'sport', '-sexo')

            else:
                context['teams'] = Team.objects.filter(
                    unit=user.unit,
                    event=event
                ).order_by('name')

                context['teams_sport'] = Team_sport.objects.filter(
                    team__unit=user.unit,
                    event=event
                ).order_by('team', 'sport', '-sexo')

            context['event_sports'] = Event_sport.objects.filter(event=event)

        return render(request, 'badge.html', context)

    else:

        try:
            event = Event.objects.get(id=request.POST.get('event_data'))
        except (Event.DoesNotExist, ValueError, TypeError):
            messages.error(request, "Evento não encontrado.")
            return redirect_badge()

        # Admin acessa qualquer evento.
        # Coordenador e usuário comum acessam apenas o próprio evento.
        if not (user.is_staff or user.type == 0):
            if event != user.event_user:
                messages.error(request, "Você não tem permissão para acessar este evento.")
                return redirect_badge()

        if 'team_sport_in' in request.POST:

            players = Player_team_sport.objects.filter(
                team_sport__id=request.POST.get('team_sport_in'),
                team_sport__event=event
            )

            if user.type == 2 and not user.is_staff:
                players = players.filter(team_sport__team__unit=user.unit)

            team_sport_badge = get_object_or_404(
                Team_sport,
                id=request.POST.get('team_sport_in'),
                event=event
            )

            if user.type == 2 and not user.is_staff and team_sport_badge.team.unit != user.unit:
                messages.error(request, "Você não tem permissão para gerar este crachá.")
                return redirect_badge()

            if len(players) == 0:
                messages.error(request, "Não tem nenhum atleta cadastrado!")
            else:
                namebadge = f'{team_sport_badge.sport.get_sport_display()}-{team_sport_badge.team.name}-jifs'
                return generate_badges(players, '7', namebadge, event.id)

        elif 'team_in' in request.POST:
            
            team = get_object_or_404(Team, id=request.POST.get('team_in'))

            if not _acesso_team(request.user, team):
                messages.error(request, "Você não tem permissão para gerar crachás deste time.")
                return redirect('badge')
            
            players_qs = Player_team_sport.objects.filter(
                team_sport__team=team,
                team_sport__event=event
            )

            seen_players = set()
            players = []

            for pts in players_qs:
                if pts.player_id not in seen_players:
                    players.append(pts)
                    seen_players.add(pts.player_id)

            if len(players) == 0:
                messages.error(request, "Não tem nenhum atleta cadastrado!")
            else:
                namebadge = f'{team.name}-jifs'
                return generate_badges(players, '7', namebadge, event.id)

        elif 'all_voluntary' in request.POST:
            voluntary = voluntary_queryset(event, 0)

            if len(voluntary) == 0:
                messages.error(request, "Não tem nenhum voluntário cadastrado!")
            else:
                namebadge = 'voluntarios-jifs'
                return generate_badges(voluntary, '0', namebadge, event.id)

        elif 'all_technician' in request.POST:
            voluntary = voluntary_queryset(event, 1)

            if len(voluntary) == 0:
                messages.error(request, "Não tem nenhum técnico cadastrado!")
            else:
                namebadge = 'tecnico-modalidade-jifs'
                return generate_badges(voluntary, '1', namebadge, event.id)

        elif 'all_support' in request.POST:
            voluntary = voluntary_queryset(event, 2)

            if len(voluntary) == 0:
                messages.error(request, "Não tem nenhum membro do apoio cadastrado!")
            else:
                namebadge = 'apoio-jifs'
                return generate_badges(voluntary, '2', namebadge, event.id)

        elif 'all_trainee' in request.POST:
            voluntary = voluntary_queryset(event, 3)

            if len(voluntary) == 0:
                messages.error(request, "Não tem nenhum estagiário cadastrado!")
            else:
                namebadge = 'estagiario-jifs'
                return generate_badges(voluntary, '3', namebadge, event.id)

        elif 'all_head' in request.POST:
            voluntary = voluntary_queryset(event, 4)

            if len(voluntary) == 0:
                messages.error(request, "Não tem nenhum chefe de delegação cadastrado!")
            else:
                namebadge = 'chefe-delegacao-jifs'
                return generate_badges(voluntary, '4', namebadge, event.id)

        elif 'all_organization' in request.POST:
            voluntary = voluntary_queryset(event, 5)

            if len(voluntary) == 0:
                messages.error(request, "Não tem nenhum membro da organização cadastrado!")
            else:
                namebadge = 'organização-jifs'
                return generate_badges(voluntary, '5', namebadge, event.id)

        elif 'all_arbitrator' in request.POST:
            voluntary = voluntary_queryset(event, 6)

            if len(voluntary) == 0:
                messages.error(request, "Não tem nenhum árbitro cadastrado!")
            else:
                namebadge = 'arbitragem-jifs'
                return generate_badges(voluntary, '6', namebadge, event.id)

        elif 'all_commission' in request.POST:
            voluntary = voluntary_queryset(event)

            if len(voluntary) == 0:
                messages.error(request, "Não tem ninguém da comissão técnica cadastrado!")
            else:
                namebadge = 'comissao-tecnica-jifs'
                return generate_badges(voluntary, None, namebadge, event.id)

    return redirect_badge()

@login_required(login_url="login")
@terms_accept_required
@permission_required('app.add_certificate', raise_exception=True)
@permission_required('app.view_certificate', raise_exception=True)
def generator_certificate(request):
    try:
        user = User.objects.get(id=request.user.id)
        if request.user.is_staff:
            team_sport = Team_sport.objects.all()
        else:
            team_sport = Team_sport.objects.filter(admin__id=request.user.id).order_by('team','sport','-sexo')
        certificate = Certificate.objects.filter(user=request.user.id)
        sport = Sport_types.choices
        if request.method == "GET":
            context = {
                'team_sport': team_sport,
                'sport': sport,
                'certificate': certificate,
                
            }
            return render(request, 'certificate.html', context)
        else:
            if 'certificate_delete' in request.POST:
                certificate_delete = request.POST.get('certificate_delete')
                certificate = Certificate.objects.get(id=certificate_delete)
                certificate.file.delete()
                certificate.delete()
                return redirect('certificate')
            elif 'certificate_all_delete' in request.POST:
                certificate = Certificate.objects.all()
                for i in certificate:
                    i.file.delete()
                    i.delete()
                return redirect('certificate')
            elif 'team-certificate' in request.POST:
                team_certificate = request.POST.get('team-certificate')
                if team_certificate.isdigit(): 
                    team_sport = get_object_or_404(Team_sport, id=team_certificate) 
                    players = Player_team_sport.objects.filter(team_sport=team_sport)
                    generate_certificates(players, user, '2')
                else:
                    if team_certificate == 'all_player':
                        players = Player_team_sport.objects.all()
                        generate_certificates(players, user, 0)
                    elif team_certificate == 'all_voluntary':
                        voluntary = Voluntary.objects.filter(type_voluntary=0, admin=user)
                        generate_certificates(voluntary, user, 1)
                    elif team_certificate == 'all_organization':
                        voluntary = Voluntary.objects.filter(type_voluntary=1, admin=user)
                        generate_certificates(voluntary, user, 1)
                    elif team_certificate == 'all_technician':
                        voluntary = Voluntary.objects.filter(type_voluntary=2, admin=user)
                        generate_certificates(voluntary, user, 1)
                    else:
                        for choice in Sport_types.choices:
                            if choice[1] == team_certificate:
                                sport_value = choice[0]
                                break
                        players = Player_team_sport.objects.filter(team_sport__sport=sport_value)
                return redirect('certificate')
            return redirect('certificate')
    except Exception as e:
        messages.error(request, f'Um erro inesperado aconteceu: {str(e)}')
    return redirect('certificate')

@login_required(login_url="login")
@terms_accept_required
def generator_data(request):
    user = User.objects.get(id=request.user.id)
    current_get_params = request.GET.urlencode()

    def _get_event():
        if user.type == 0:
            event_id = request.GET.get('e') or request.POST.get('event_data')
            if event_id:
                return Event.objects.filter(id=event_id).first()
            return None
        return user.event_user

    def _base_team_sports(event):
        qs = (
            Team_sport.objects
            .filter(event=event, players__isnull=False)
            .select_related('team', 'sport', 'event')
            .prefetch_related(
                Prefetch(
                    'players',
                    queryset=Player_team_sport.objects.select_related('player')
                )
            )
            .distinct()
        )

        if user.type == 2:
            if user.team:
                qs = qs.filter(team=user.team)
            elif user.unit:
                qs = qs.filter(team__unit=user.unit)

        return qs.order_by('sport__sport', 'team__name', '-sexo')

    def _build_modalidades_relatorio(event, base_qs):
        modalidades = []

        for es in Event_sport.objects.filter(event=event).order_by('sport'):
            qs_es = base_qs.filter(sport=es)

            if not qs_es.exists():
                continue

            modalidades.append({
                'event_sport': es,
                'has_full': qs_es.exists(),
                'has_masc': qs_es.filter(sexo=0).exists(),
                'has_fem': qs_es.filter(sexo=1).exists(),
            })

        return modalidades

    def _base_voluntarys(event):
        qs = Voluntary.objects.filter(event=event).select_related('unit', 'admin').order_by('name')

        if user.type == 2:
            if user.unit:
                qs = qs.filter(unit=user.unit)
            elif user.team and user.team.unit:
                qs = qs.filter(unit=user.team.unit)
            else:
                qs = Voluntary.objects.none()

        return qs

    # ── GET ───────────────────────────────────────────────────────────────────
    if request.method == "GET":
        event = _get_event()

        context = {
            'events': Event.objects.all(),
            'type_service': Type_service.choices,
        }

        if not event:
            return render(request, 'data.html', context)

        base_ts = _base_team_sports(event)

        if user.type == 2:
            teams = Team.objects.filter(id=user.team.id) if user.team else Team.objects.none()
        else:
            teams = Team.objects.filter(event=event).order_by('name')

        # tipos disponíveis de voluntários no evento (já respeitando acesso)
        available_voluntary_types = (
            _base_voluntarys(event)
            .values_list('type_voluntary', flat=True)
            .distinct()
        )

        context.update({
            'event': event,
            'teams': teams,
            'my_team': user.team if user.type == 2 else None,
            'modalidades_relatorio': _build_modalidades_relatorio(event, base_ts),
            'available_voluntary_types': list(available_voluntary_types),
        })

        return render(request, 'data.html', context)

    # ── POST ──────────────────────────────────────────────────────────────────
    event = _get_event()
    if not event:
        messages.error(request, "Nenhum evento selecionado.")
        return redirect('data')

    status = False

    cont = {
        'now': timezone.now(),
        'user': request.user,
        'event': event,
        'logo_morea': request.build_absolute_uri('/static/images/logo_atum.png')
    }

    if event:
        cont['logo_ifs'] = request.build_absolute_uri(event.logo.url)
        cont['event'] = event
    else:
        cont['logo_ifs'] = request.build_absolute_uri('/static/images/logo-jiifs-2025.jpg')

    logo_ifs = cont['logo_ifs']
    img = ImageReader(logo_ifs)
    largura, altura = img.getSize()

    if largura == altura:
        cont['logo_event_type'] = 0
    elif largura > altura:
        cont['logo_event_type'] = 1
    else:
        cont['logo_event_type'] = 2

    base_ts = _base_team_sports(event)

    # ─────────────────────────────────────────────────────────────────────────
    # DADOS GERAIS
    # ─────────────────────────────────────────────────────────────────────────
    if 'all_data' in request.POST:
        if user.type not in (0, 1) and not user.is_staff:
            messages.error(request, "Sem permissão para este relatório.")
            return redirect('data')

        name_html = 'data-general'
        name_pdf = 'dados_gerais'

        qnt_players = Player.objects.filter(event=event).count()
        qnt_players_fem = Player.objects.filter(sexo=1, event=event).count()
        qnt_players_masc = Player.objects.filter(sexo=0, event=event).count()

        cont['qnt_players'] = qnt_players
        cont['qnt_players_fem'] = qnt_players_fem
        cont['qnt_players_masc'] = qnt_players_masc
        cont['qnt_teams'] = Team_sport.objects.filter(event=event).count()
        cont['qnt_voluntary_0'] = Voluntary.objects.filter(type_voluntary=0, event=event).count()
        cont['qnt_voluntary_1'] = Voluntary.objects.filter(type_voluntary=1, event=event).count()
        cont['qnt_voluntary_2'] = Voluntary.objects.filter(type_voluntary=2, event=event).count()
        cont['qnt_voluntary_3'] = Voluntary.objects.filter(type_voluntary=3, event=event).count()
        cont['qnt_voluntary_4'] = Voluntary.objects.filter(type_voluntary=4, event=event).count()
        cont['qnt_voluntary_5'] = Voluntary.objects.filter(type_voluntary=5, event=event).count()
        cont['qnt_voluntary_6'] = Voluntary.objects.filter(type_voluntary=6, event=event).count()

        if qnt_players > 0:
            cont['porcent_fem'] = (qnt_players_fem * 100) / qnt_players
            cont['porcent_masc'] = (qnt_players_masc * 100) / qnt_players
        else:
            cont['porcent_fem'] = 0
            cont['porcent_masc'] = 0

    # ─────────────────────────────────────────────────────────────────────────
    # INSCRIÇÕES
    # ─────────────────────────────────────────────────────────────────────────
    elif 'enrollment' in request.POST:
        if user.type not in (0, 1) and not user.is_staff:
            messages.error(request, "Sem permissão para este relatório.")
            return redirect('data')

        name_html = 'data-base-enrollment'
        name_pdf = 'relatório_de_inscrições'

        teams = (
            Team_sport.objects
            .prefetch_related('players')
            .filter(team__event=event)
            .order_by('team', 'sport', '-sexo')
        )

        if not teams.exists():
            messages.error(request, "Não há equipes em modalidades ou atletas cadastrados.")
            status = True

        cont['teams'] = teams

    # ─────────────────────────────────────────────────────────────────────────
    # DADOS DOS TIMES
    # ─────────────────────────────────────────────────────────────────────────
    elif 'all_team' in request.POST:
        name_html = 'data-base-campus'
        name_pdf = 'dados_campus'

        teams = base_ts

        if not teams.exists():
            messages.error(request, "Não há equipes ou atletas cadastrados.")
            status = True

        cont['teams'] = teams
        cont['infor'] = "campus x modalidade x atletas"

    # ─────────────────────────────────────────────────────────────────────────
    # PARTIDAS
    # ─────────────────────────────────────────────────────────────────────────
    elif 'all_match' in request.POST:
        if user.type not in (0, 1) and not user.is_staff:
            messages.error(request, "Sem permissão para este relatório.")
            return redirect('data')

        name_html = 'data-base-match'
        name_pdf = 'relatório_de_partidas'

        matches = Match.objects.filter(event=event).prefetch_related('teams__team').order_by('time_match')

        if not matches.exists():
            messages.error(request, "Não há partidas cadastradas.")
            status = True

        cont['matches'] = matches

    # ─────────────────────────────────────────────────────────────────────────
    # TODOS OS ATLETAS
    # ─────────────────────────────────────────────────────────────────────────
    elif 'all_players' in request.POST:
        name_html = 'data-base'
        name_pdf = 'dados_todos_atletas'

        if user.is_staff or user.type in (0, 1):
            players = Player.objects.filter(event=event).order_by('name')
        else:
            if user.unit:
                players = Player.objects.filter(unit=user.unit, event=event).order_by('name')
            elif user.team:
                players = Player.objects.filter(
                    player_team_sport__team_sport__team=user.team,
                    event=event
                ).distinct().order_by('name')
            else:
                players = Player.objects.none()

        if not players.exists():
            messages.error(request, "Não há atletas cadastrados.")
            status = True

        cont['players'] = players
        cont['infor'] = "todos os atletas"
        cont['type'] = True

    # ─────────────────────────────────────────────────────────────────────────
    # ATLETAS FEMININAS
    # ─────────────────────────────────────────────────────────────────────────
    elif 'all_players_fem' in request.POST:
        name_html = 'data-base'
        name_pdf = 'dados_atletas_femininas'

        if user.is_staff or user.type in (0, 1):
            players = Player.objects.filter(sexo=1, event=event).order_by('name')
        else:
            if user.unit:
                players = Player.objects.filter(sexo=1, unit=user.unit, event=event).order_by('name')
            elif user.team:
                players = Player.objects.filter(
                    sexo=1,
                    player_team_sport__team_sport__team=user.team,
                    event=event
                ).distinct().order_by('name')
            else:
                players = Player.objects.none()

        if not players.exists():
            messages.error(request, "Não há atletas do sexo feminino cadastradas.")
            status = True

        cont['players'] = players
        cont['infor'] = "atletas do sexo feminino"
        cont['type'] = True

    # ─────────────────────────────────────────────────────────────────────────
    # ATLETAS MASCULINOS
    # ─────────────────────────────────────────────────────────────────────────
    elif 'all_players_masc' in request.POST:
        name_html = 'data-base'
        name_pdf = 'dados_atletas_masculinos'

        if user.is_staff or user.type in (0, 1):
            players = Player.objects.filter(sexo=0, event=event).order_by('name')
        else:
            if user.unit:
                players = Player.objects.filter(sexo=0, unit=user.unit, event=event).order_by('name')
            elif user.team:
                players = Player.objects.filter(
                    sexo=0,
                    player_team_sport__team_sport__team=user.team,
                    event=event
                ).distinct().order_by('name')
            else:
                players = Player.objects.none()

        if not players.exists():
            messages.error(request, "Não há atletas do sexo masculino cadastrados.")
            status = True

        cont['players'] = players
        cont['infor'] = "atletas do sexo masculino"
        cont['type'] = True

    # ─────────────────────────────────────────────────────────────────────────
    # TODOS OS VOLUNTÁRIOS / COMISSÃO / APOIO / ETC
    # ─────────────────────────────────────────────────────────────────────────
    elif 'all_eqp' in request.POST:
        name_html = 'data-base-eqp'
        name_pdf = 'comissao_tecnica'

        voluntarys = _base_voluntarys(event)

        if not voluntarys.exists():
            messages.error(request, "Não há membros cadastrados.")
            status = True

        cont['voluntarys'] = voluntarys
        cont['infor'] = "todos os membros"

    # ─────────────────────────────────────────────────────────────────────────
    # VOLUNTÁRIOS POR TIPO
    # ─────────────────────────────────────────────────────────────────────────
    elif 'all_eqp_type' in request.POST:
        type_value = request.POST.get('all_eqp_type')

        try:
            type_value = int(type_value)
        except (TypeError, ValueError):
            messages.error(request, "Tipo de membro inválido.")
            return redirect('data')

        name_html = 'data-base-eqp'
        type_label = dict(Type_service.choices).get(type_value, 'membros')
        name_pdf = f'relatorio_{type_label.lower()}'

        voluntarys = _base_voluntarys(event).filter(type_voluntary=type_value)

        if not voluntarys.exists():
            messages.error(request, "Não há membros cadastrados nesse tipo.")
            status = True

        cont['voluntarys'] = voluntarys

        type_map = dict(Type_service.choices)
        cont['infor'] = type_map.get(type_value, "membros")

    # ─────────────────────────────────────────────────────────────────────────
    # CAMPUS INDIVIDUAL
    # ─────────────────────────────────────────────────────────────────────────
    elif 'team_in' in request.POST:
        if user.type not in (0, 1) and not user.is_staff:
            messages.error(request, "Sem permissão para este relatório.")
            return redirect('data')

        team_id = request.POST.get('team_in')
        try:
            team = Team.objects.get(id=team_id, event=event)
        except Team.DoesNotExist:
            messages.error(request, "Campus não encontrado neste evento.")
            return redirect('data')

        name_html = 'data-base-campus-individual'
        name_pdf = f'atletas_{team.name}'

        players = (
            Player_team_sport.objects
            .filter(team_sport__team=team, player__event=event)
            .select_related('player', 'team_sport__team', 'team_sport__sport')
            .order_by('player__name')
        )

        if not players.exists():
            messages.error(request, "Não há atletas cadastrados neste campus.")
            status = True

        cont['team'] = team
        cont['players'] = players
        cont['infor'] = "atletas"
        cont['campus'] = team.name

    # ─────────────────────────────────────────────────────────────────────────
    # MODALIDADE COMPLETA
    # ─────────────────────────────────────────────────────────────────────────
    elif 'all_players_sport' in request.POST:
        data = request.POST.get('all_players_sport')
        event_sport = get_object_or_404(Event_sport, id=data, event=event)

        name_html = 'data-base-campus'
        name_pdf = f'atletas_{event_sport.get_sport_display()}'

        teams = base_ts.filter(sport=event_sport)

        if not teams.exists():
            messages.error(request, "Não há equipes ou atletas cadastrados nessa modalidade.")
            status = True

        cont['teams'] = teams
        cont['infor'] = event_sport.get_sport_display()

    # ─────────────────────────────────────────────────────────────────────────
    # MODALIDADE MASCULINA
    # ─────────────────────────────────────────────────────────────────────────
    elif 'all_players_sport_masc' in request.POST:
        data = request.POST.get('all_players_sport_masc')
        event_sport = get_object_or_404(Event_sport, id=data, event=event)

        name_html = 'data-base-campus'
        name_pdf = f'atletas_{event_sport.get_sport_display()}_masculino'

        teams = base_ts.filter(sport=event_sport, sexo=0)

        if not teams.exists():
            messages.error(request, "Não há equipes ou atletas masculinos cadastrados nessa modalidade.")
            status = True

        cont['teams'] = teams
        cont['infor'] = f"{event_sport.get_sport_display()} - Masculino"

    # ─────────────────────────────────────────────────────────────────────────
    # MODALIDADE FEMININA
    # ─────────────────────────────────────────────────────────────────────────
    elif 'all_players_sport_fem' in request.POST:
        data = request.POST.get('all_players_sport_fem')
        event_sport = get_object_or_404(Event_sport, id=data, event=event)

        name_html = 'data-base-campus'
        name_pdf = f'atletas_{event_sport.get_sport_display()}_feminino'

        teams = base_ts.filter(sport=event_sport, sexo=1)

        if not teams.exists():
            messages.error(request, "Não há equipes ou atletas femininas cadastradas nessa modalidade.")
            status = True

        cont['teams'] = teams
        cont['infor'] = f"{event_sport.get_sport_display()} - Feminino"

    else:
        messages.error(request, "Nenhum relatório foi selecionado.")
        return redirect('data')

    if status:
        if current_get_params:
            return redirect(f"{reverse('data')}?{current_get_params}")
        return redirect('data')

    html_string = render_to_string(f'generator/{name_html}.html', cont)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{name_pdf}.pdf"'
    HTML(string=html_string).write_pdf(response)

    return response

@login_required(login_url="login")
@terms_accept_required
def generator_spreadsheet(request):
    user = request.user

    # permite:
    # - quem tem a permissão app.data
    # - técnico (type == 2)
    if not (user.has_perm('app.data') or user.type in [1, 2]):
        raise PermissionDenied

    current_get_params = request.GET.urlencode()

    def _redirect():
        if current_get_params:
            return redirect(f"{reverse('spreadsheet')}?{current_get_params}")
        return redirect('spreadsheet')

    def _dedup_from_pts(pts_qs):
        seen = set()
        result = []
        for pts in pts_qs:
            if pts.player_id not in seen:
                result.append(pts.player)
                seen.add(pts.player_id)
        return result

    # GET — monta contexto
    if request.method == "GET":
        context = {}

        if user.type == 0:
            context['events'] = Event.objects.all()

            selected_event = request.GET.get('e')
            if selected_event:
                try:
                    event = Event.objects.get(id=selected_event)
                    context['event'] = event
                    context['select_event'] = selected_event
                    context['teams'] = Team.objects.filter(event=event).order_by('name')
                    context['event_sports'] = Event_sport.objects.filter(event=event)
                except Event.DoesNotExist:
                    messages.error(request, "Evento selecionado não foi encontrado.")

        elif user.type == 1:
            event = user.event_user
            if event:
                context['event'] = event
                context['teams'] = Team.objects.filter(event=event).order_by('name')
                context['event_sports'] = Event_sport.objects.filter(event=event)
            else:
                messages.error(request, "Seu usuário não possui evento vinculado.")

        elif user.type == 2:
            if user.event_user and user.team:
                event = user.event_user
                context['event'] = event
                context['my_team'] = user.team
                context['my_team_sports'] = (
                    Team_sport.objects
                    .filter(team=user.team, event=event)
                    .select_related('sport')
                    .order_by('sport__sport', 'sexo')
                )
            else:
                messages.error(request, "Seu usuário técnico precisa ter evento e campus vinculados.")

        return render(request, 'spreadsheet.html', context)

    # POST — gera o xlsx
    event_id = request.POST.get('event_data')
    if not event_id:
        messages.error(request, "Evento não identificado.")
        return _redirect()

    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        messages.error(request, "Evento não encontrado.")
        return _redirect()

    # segurança extra por tipo de usuário
    if user.type in [1, 2] and user.event_user != event:
        messages.error(request, "Você não tem permissão para acessar este evento.")
        return _redirect()

    players = None
    label = event.name

    # GERAL
    if 'p_todos' in request.POST:
        if user.type == 2:
            players = list(
                Player.objects.filter(event=event, admin=user).order_by('name')
            )
        else:
            players = list(
                Player.objects.filter(event=event).order_by('name')
            )
        label = f'todos_{event.name}'

    elif 'p_masc' in request.POST:
        if user.type == 2:
            players = list(
                Player.objects.filter(event=event, admin=user, sexo=0).order_by('name')
            )
        else:
            players = list(
                Player.objects.filter(event=event, sexo=0).order_by('name')
            )
        label = f'masculino_{event.name}'

    elif 'p_fem' in request.POST:
        if user.type == 2:
            players = list(
                Player.objects.filter(event=event, admin=user, sexo=1).order_by('name')
            )
        else:
            players = list(
                Player.objects.filter(event=event, sexo=1).order_by('name')
            )
        label = f'feminino_{event.name}'

    # POR CAMPUS
    elif 'p_team' in request.POST:
        if user.type not in [0, 1]:
            messages.error(request, "Você não tem permissão para gerar esta planilha.")
            return _redirect()

        try:
            team = Team.objects.get(id=request.POST.get('p_team'), event=event)
        except Team.DoesNotExist:
            messages.error(request, "Campus não encontrado neste evento.")
            return _redirect()

        pts_qs = (
            Player_team_sport.objects
            .filter(team_sport__team=team, player__event=event)
            .select_related('player')
            .order_by('player__name')
        )
        players = _dedup_from_pts(pts_qs)
        label = f'{team.name}_{event.name}'

    # POR MODALIDADE — superusuário/coordenador
    elif 'p_sport' in request.POST:
        if user.type not in [0, 1]:
            messages.error(request, "Você não tem permissão para gerar esta planilha.")
            return _redirect()

        try:
            event_sport = Event_sport.objects.get(id=request.POST.get('p_sport'), event=event)
        except Event_sport.DoesNotExist:
            messages.error(request, "Modalidade não encontrada neste evento.")
            return _redirect()

        pts_qs = (
            Player_team_sport.objects
            .filter(team_sport__sport=event_sport, player__event=event)
            .select_related('player')
            .order_by('player__name')
        )
        players = _dedup_from_pts(pts_qs)
        label = f'{event_sport.get_sport_display()}_{event.name}'

    # POR MODALIDADE + GÊNERO — técnico
    elif 'p_team_sport' in request.POST:
        try:
            team_sport_obj = Team_sport.objects.select_related('team', 'sport').get(
                id=request.POST.get('p_team_sport'),
                event=event
            )
        except Team_sport.DoesNotExist:
            messages.error(request, "Modalidade da equipe não encontrada.")
            return _redirect()

        if user.type == 2 and team_sport_obj.team != user.team:
            messages.error(request, "Você não tem permissão para gerar esta planilha.")
            return _redirect()

        pts_qs = (
            Player_team_sport.objects
            .filter(team_sport=team_sport_obj)
            .select_related('player')
            .order_by('player__name')
        )
        players = _dedup_from_pts(pts_qs)
        label = (
            f'{team_sport_obj.sport.get_sport_display()}_'
            f'{team_sport_obj.get_sexo_display()}_'
            f'{event.name}'
        )

    else:
        messages.error(request, "Nenhum filtro de planilha foi enviado.")
        return _redirect()

    if not players:
        messages.error(request, "Não há atletas cadastrados para os filtros selecionados.")
        return _redirect()

    return generate_spreadsheet(players, event.name, label)


# Dashboard de acesso

def generate_spreadsheet(players, event_name, label):
    wb = Workbook()

    thin = Side(style='thin', color='000000')
    medium = Side(style='medium', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font = Font(name='Arial', size=10, bold=True)
    body_font = Font(name='Arial', size=10)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    fill_header_pag = PatternFill(fill_type='solid', fgColor='BDBDBD')
    fill_header_end = PatternFill(fill_type='solid', fgColor='AEABAB')
    fill_white = PatternFill(fill_type='solid', fgColor='FFFFFF')

    def make_header(ws, headers, widths, fill_color):
        ws.row_dimensions[1].height = 35

        for i, (header, width) in enumerate(zip(headers, widths), start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
            cell = ws.cell(row=1, column=i, value=header)
            cell.font = header_font
            cell.fill = PatternFill(fill_type='solid', fgColor=fill_color)
            cell.alignment = center
            cell.border = Border(left=thin, right=thin, top=medium, bottom=thin)

    def style_row(ws, row_num, num_cols, center_cols=()):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = body_font
            cell.fill = fill_white
            cell.border = border
            cell.alignment = center if col in center_cols else left

    # ABA 1 - PAGAMENTO
    ws1 = wb.active
    ws1.title = 'PLANILHA DE PAGAMENTO'

    headers_pag = [
        'SEQ',
        'MATRÍCULA',
        'NOME DO ALUNO',
        'CPF/CHAVE PIX [Somente Números]',
        'VALOR',
        'CURSO',
        'AUXÍLIO [Descrição]',
        'EVENTO',
        'VALOR TOTAL',
    ]

    widths_pag = [5.86, 13.29, 44.14, 24.00, 14.29, 29.71, 20.00, 20.00, 16.00]
    make_header(ws1, headers_pag, widths_pag, 'BDBDBD')

    for seq, player in enumerate(players, start=1):
        ws1.append([
            seq,
            player.registration or '',
            player.name or '',
            (player.cpf or '').replace('.', '').replace('-', '').replace('/', ''),
            '',          # VALOR
            player.course or '',
            '',          # AUXÍLIO [Descrição]
            event_name,  # EVENTO
            '',          # VALOR TOTAL fica vazio nas linhas
        ])

        style_row(ws1, ws1.max_row, len(headers_pag), center_cols=(1, 2, 4, 5, 8, 9))

    # linha de total geral
    first_data_row = 2
    last_data_row = ws1.max_row
    total_row = last_data_row + 1

    # deixa a linha inteira estilizada
    style_row(ws1, total_row, len(headers_pag), center_cols=(1, 2, 4, 5, 8, 9))

    # rótulo do total
    ws1.cell(row=total_row, column=8, value='TOTAL GERAL')
    ws1.cell(row=total_row, column=8).font = header_font
    ws1.cell(row=total_row, column=8).alignment = center
    ws1.cell(row=total_row, column=8).border = border
    ws1.cell(row=total_row, column=8).fill = fill_white

    # fórmula somando toda a coluna VALOR
    if last_data_row >= first_data_row:
        ws1.cell(row=total_row, column=9, value=f'=SUM(E{first_data_row}:E{last_data_row})')
    else:
        ws1.cell(row=total_row, column=9, value='')

    ws1.cell(row=total_row, column=9).font = header_font
    ws1.cell(row=total_row, column=9).alignment = center
    ws1.cell(row=total_row, column=9).border = border
    ws1.cell(row=total_row, column=9).fill = fill_white

    for row in range(2, ws1.max_row + 1):
        ws1.cell(row=row, column=5).number_format = 'R$ #,##0.00'
        ws1.cell(row=row, column=9).number_format = 'R$ #,##0.00'

    # ABA 2 - ENDEREÇO
    ws2 = wb.create_sheet('PLANILHA DE ENDEREÇO')

    headers_end = [
        'SEQ.',
        'MATRÍCULA',
        'NOME DO ALUNO',
        'CPF / CHAVE PIX [Somente Números]',
        'CURSO',
        'ENDEREÇO COMPLETO [Rua, Bairro e Número]',
        'CEP [Somente Números]',
        'MUNICÍPIO/UF',
    ]

    widths_end = [5.86, 13.29, 37.57, 24.00, 29.71, 63.86, 18.00, 26.0]
    make_header(ws2, headers_end, widths_end, 'AEABAB')

    for seq, player in enumerate(players, start=1):
        ws2.append([
            seq,
            player.registration or '',
            player.name or '',
            (player.cpf or '').replace('.', '').replace('-', '').replace('/', ''),
            player.course or '',
            player.address or '',
            (player.cep or '').replace('-', '').replace('.', ''),
            player.municipality or '',
        ])

        style_row(ws2, ws2.max_row, len(headers_end), center_cols=(1, 2, 4, 7, 8))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_label = (
        label.replace(' ', '_')
        .replace('/', '_')
        .replace('\\', '_')
        .lower()
    )

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="planilha_{safe_label}.xlsx"'
    return response
